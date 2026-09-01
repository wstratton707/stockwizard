import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from datetime import datetime

DARK_BLUE = "1F4E79"
MID_BLUE  = "2E75B6"
GREEN_OK  = "70AD47"
RED_BAD   = "FF0000"
AMBER     = "FFC000"
WHITE     = "FFFFFF"
GREY_ROW  = "F2F2F2"
LIGHT_BG  = "EBF5FB"


GRID_GREY = "D9D9D9"


def _border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _soft_line(el):
    """Hairline grey — for gridlines and axis rules."""
    el.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID_GREY, w=9525))
    return el


def _chart_chrome(ch, bar_colour=None, num_fmt=None, gridlines=True):
    """Make an openpyxl chart legible.

    Two Excel defaults ruin these on sight and neither is visible from the code:

    `delete` is unset on both axes, and Excel reads that as "hide" — so the
    sector and position charts shipped with no category names and no value
    scale. Bars with nothing to identify them are decoration, not a chart.

    `varyColors` defaults on for single-series bar charts, so every bar took a
    different colour. Colour that encodes nothing invites the reader to look for
    a meaning that isn't there; one series should be one colour.

    Rounded corners, a black frame and near-black gridlines are the rest of the
    untouched-default look.
    """
    ch.roundedCorners = False
    ch.varyColors = False
    # Titles default to overlaying the plot, which on the tall Position Weights
    # chart printed the title across its own top bar.
    try:
        if ch.title is not None:
            ch.title.overlay = False
    except Exception:
        pass
    ch.graphical_properties = GraphicalProperties()
    ch.graphical_properties.line.noFill = True
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    _soft_line(ch.x_axis)
    _soft_line(ch.y_axis)
    if gridlines:
        ch.y_axis.majorGridlines = _soft_line(ChartLines())
    if num_fmt:
        ch.y_axis.numFmt = num_fmt
    if ch.legend is not None:
        ch.legend.position = "b"
        ch.legend.overlay = False
    if bar_colour:
        for ser in ch.series:
            ser.graphicalProperties.solidFill = bar_colour
            ser.graphicalProperties.line.noFill = True
    return ch


def _hdr(cell, bg=DARK_BLUE, fg=WHITE, size=10):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


def _kv(ws, row, label, value, fmt=None, rag=None, label_col=1, val_col=2):
    cl = ws.cell(row=row, column=label_col, value=label)
    cv = ws.cell(row=row, column=val_col,   value=value)
    cl.font      = Font(name="Arial", size=10)
    cv.font      = Font(name="Arial", size=10, bold=True)
    cv.alignment = Alignment(horizontal="right")
    cl.border    = cv.border = _border()
    if fmt and isinstance(value, (int, float)):
        cv.number_format = fmt
    if rag and isinstance(value, (int, float)):
        d, thresh = rag
        colour = (GREEN_OK if value > thresh else RED_BAD) if d == "gt" \
                 else (RED_BAD if value < thresh else GREEN_OK)
        cv.fill = PatternFill("solid", fgColor=colour)
        cv.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    bg = GREY_ROW if row % 2 == 0 else WHITE
    for c in [label_col, val_col]:
        cell = ws.cell(row=row, column=c)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000","FFFFFFFF",WHITE):
            cell.fill = PatternFill("solid", fgColor=bg)


def _sec_hdr(ws, row, label, col_start=1, col_end=4):
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
    c = ws.cell(row=row, column=col_start, value=label)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=11)
    c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def _auto_width(ws, max_w=30):
    from openpyxl.cell.cell import Cell
    for col in ws.columns:
        real = [c for c in col if isinstance(c, Cell)]
        if not real:
            continue
        best = max((len(str(c.value or "")) for c in real), default=10)
        ws.column_dimensions[real[0].column_letter].width = min(best + 3, max_w)


def _style_header_row(ws, bg=DARK_BLUE):
    for cell in ws[1]:
        _hdr(cell, bg=bg)


# ── Cover ─────────────────────────────────────────────────────────────────────
def _build_cover(wb, preferences, final_weights, backtest_metrics, mc_summary, sheetnames):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 28

    ws.merge_cells("B2:C3")
    c = ws["B2"]
    c.value     = "◈  QuantWizard — Portfolio Analysis Report"
    c.font      = Font(size=20, bold=True, color=MID_BLUE, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B4:C4")
    ws["B4"].value = (f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
                      f"  |  Multi-source data: Polygon · Yahoo Finance · Finnhub · SEC EDGAR")
    ws["B4"].font  = Font(size=9, italic=True, color="888888", name="Arial")

    ws["B6"] = "TABLE OF CONTENTS"
    ws["B6"].font = Font(bold=True, size=12, color=DARK_BLUE, name="Arial")
    for i, name in enumerate([s for s in sheetnames if s != "Cover"], 7):
        cell = ws.cell(row=i, column=2, value=name.replace("_", " "))
        cell.font      = Font(name="Arial", size=10, color=MID_BLUE, underline="single")
        cell.hyperlink = f"#{name}!A1"
        ws.row_dimensions[i].height = 16

    # Key stats block
    snap_row = 7 + len(sheetnames) + 2
    stats = [
        ("Risk Tolerance",      f"{preferences.get('risk_tolerance',5)}/10"),
        ("Investment Horizon",  preferences.get("horizon","5 years")),
        ("Starting Capital",    f"${preferences.get('starting_capital',10000):,.0f}"),
        ("Monthly Contribution",f"${preferences.get('monthly_contribution',500):,.0f}"),
        ("Holdings",            len(final_weights)),
        ("Final Portfolio Value", f"${backtest_metrics.get('Final Value',0):,.2f}"),
        ("Total Return",        f"{backtest_metrics.get('Total Return',0):.2f}%"),
        ("Sharpe Ratio",        (f"{backtest_metrics['Sharpe Ratio']:.2f}"
                                 if isinstance(backtest_metrics.get("Sharpe Ratio"), (int, float))
                                 else "N/A")),
    ]
    ws.cell(row=snap_row, column=2, value="Portfolio Snapshot").font = Font(
        bold=True, size=11, color=DARK_BLUE, name="Arial")
    for j, (k, v) in enumerate(stats, snap_row+1):
        ws.cell(row=j, column=2, value=k).font  = Font(name="Arial", size=9, color="555555")
        ws.cell(row=j, column=3, value=str(v)).font = Font(name="Arial", size=9, bold=True)

    ws["B2"].fill = PatternFill("solid", fgColor="EBF5FB")


# ── Dashboard ─────────────────────────────────────────────────────────────────
def _build_dashboard(wb, preferences, final_weights, stock_metrics,
                     backtest_metrics, mc_summary, diversification_score):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10

    ws.merge_cells("A1:D1")
    ws["A1"] = "◈ QuantWizard — Portfolio Dashboard"
    ws["A1"].font      = Font(size=18, bold=True, color=DARK_BLUE, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}  |  Powered by QuantWizard"
    ws["A2"].font      = Font(italic=True, color="888888", name="Arial", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    _sec_hdr(ws, row, "Portfolio Preferences")
    row += 1
    prefs_display = [
        ("Risk Tolerance",       f"{preferences.get('risk_tolerance',5)}/10"),
        ("Investment Horizon",   preferences.get("horizon","5 years")),
        ("Starting Capital",     preferences.get("starting_capital",10000)),
        ("Monthly Contribution", preferences.get("monthly_contribution",500)),
        ("Target Goal",          preferences.get("target_value","Not set")),
    ]
    for label, val in prefs_display:
        _kv(ws, row, label, val,
            fmt='_($* #,##0.00_)' if isinstance(val,(int,float)) else None)
        row += 1

    row += 1
    _sec_hdr(ws, row, "Portfolio Performance")
    row += 1
    perf_items = [
        ("Final Portfolio Value",  backtest_metrics.get("Final Value",0),         '_($* #,##0.00_)', ("gt",0)),
        ("Total Gain / Loss",      backtest_metrics.get("Total Gain/Loss",0),      '_($* #,##0.00_)', ("gt",0)),
        ("Total Return %",         backtest_metrics.get("Total Return",0),         '0.00%',           ("gt",0)),
        # `or 0` not a .get default: the key exists but is None when the window is
        # too short to annualise, and .get would hand that None straight to openpyxl.
        ("Annualised Return %",    backtest_metrics.get("Ann. Return") or 0,       '0.00%',           ("gt",0)),
        ("vs S&P 500",             backtest_metrics.get("vs S&P 500","N/A"),       None,              None),
        # `or 0` for the same reason as Ann. Return above: the key exists and is
        # None when the window was too short to justify a Sharpe.
        ("Sharpe Ratio",           backtest_metrics.get("Sharpe Ratio") or 0,      '0.00',            ("gt",1)),
        ("Sortino Ratio",          backtest_metrics.get("Sortino Ratio",0),        '0.000',           ("gt",1)),
        ("Max Drawdown",           backtest_metrics.get("Max Drawdown",0),         '0.00%',           ("gt",-20)),
        ("Ann. Volatility",        backtest_metrics.get("Ann. Volatility",0),      '0.00%',           None),
        ("Best Month",             backtest_metrics.get("Best Month",0),           '0.00%',           ("gt",0)),
        ("Worst Month",            backtest_metrics.get("Worst Month",0),          '0.00%',           ("gt",-5)),
        ("% Months Positive",      backtest_metrics.get("% Months Positive",0),   '0.0%',            ("gt",50)),
        ("Diversification Score",  diversification_score,                          '0.0',             ("gt",6)),
    ]
    for label, val, fmt, rag in perf_items:
        if isinstance(val, (int, float)):
            # Convert percentage display values
            display_val = val/100 if fmt == '0.00%' or fmt == '0.0%' else val
            _kv(ws, row, label, display_val, fmt=fmt, rag=rag)
        else:
            _kv(ws, row, label, str(val))
        row += 1

    row += 1
    _sec_hdr(ws, row, "Monte Carlo Forecast Summary")
    row += 1
    if mc_summary:
        for k, v in mc_summary.items():
            _kv(ws, row, k, str(v))
            row += 1

    return ws


# ── Holdings sheet ────────────────────────────────────────────────────────────
def _build_holdings_sheet(wb, final_weights, stock_metrics, ticker_info):
    ws = wb.create_sheet("Holdings_Breakdown")
    headers = ["Ticker","Company","Weight %","Ann. Return %","Ann. Volatility %",
               "Sharpe Ratio","Sortino Ratio","Max Drawdown %","Total Return %"]
    ws.append(headers)
    _style_header_row(ws, bg=MID_BLUE)

    sorted_holdings = sorted(final_weights.items(), key=lambda x: x[1], reverse=True)
    for ri, (ticker, weight) in enumerate(sorted_holdings, 2):
        m    = stock_metrics.get(ticker, {})
        info = ticker_info.get(ticker, {})
        row  = [
            ticker,
            info.get("name", ticker)[:30],
            round(weight * 100, 2),
            m.get("ann_return", "N/A"),
            m.get("ann_vol",    "N/A"),
            m.get("sharpe",     "N/A"),
            m.get("sortino",    "N/A"),
            m.get("max_drawdown","N/A"),
            m.get("total_return","N/A"),
        ]
        ws.append(row)
        for ci, cell in enumerate(ws[ri], 1):
            cell.font   = Font(name="Arial", size=10)
            cell.border = _border()
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY_ROW)
            if ci == 3:  # Weight %
                cell.number_format = "0.00%"
                cell.value = weight

    # Conditional formatting on Sharpe
    sharpe_col = "D"
    ws.conditional_formatting.add(
        f"{sharpe_col}2:{sharpe_col}{ws.max_row}",
        ColorScaleRule(start_type="min", start_color="FF9999",
                       mid_type="num",  mid_value=1, mid_color="FFFFFF",
                       end_type="max",  end_color="99FF99"))

    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"


# ── Backtest sheet ────────────────────────────────────────────────────────────
def _build_backtest_sheet(wb, backtest_df, backtest_metrics):
    ws = wb.create_sheet("Backtest_Results")

    # Metrics summary at top
    ws["A1"] = "Backtest Performance Metrics"
    ws["A1"].font = Font(bold=True, size=13, color=DARK_BLUE, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"], ws["B2"] = "Metric", "Value"
    for cell in ws[2]:
        _hdr(cell, bg=MID_BLUE)

    row = 3
    for k, v in backtest_metrics.items():
        ws.cell(row=row, column=1, value=k).font  = Font(name="Arial", size=10)
        ws.cell(row=row, column=2, value=str(v)).font = Font(name="Arial", size=10, bold=True)
        row += 1

    # Daily values table
    data_start = row + 2
    ws.cell(row=data_start, column=1, value="Date").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=data_start, column=2, value="Portfolio ($)").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=data_start, column=3, value="Contributions ($)").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=data_start, column=4, value="S&P 500 ($)").font = Font(bold=True, name="Arial", size=10)
    for c in range(1, 5):
        _hdr(ws.cell(row=data_start, column=c), bg=MID_BLUE)

    # Sample every 5 days to keep file size manageable
    sample = backtest_df.iloc[::5]
    for ri, (date, row_data) in enumerate(sample.iterrows(), data_start+1):
        ws.cell(row=ri, column=1, value=date.strftime("%Y-%m-%d")).number_format = "yyyy-mm-dd"
        ws.cell(row=ri, column=2, value=round(row_data["Portfolio"],2)).number_format = '_($* #,##0.00_)'
        ws.cell(row=ri, column=3, value=round(row_data["Contrib"],2)).number_format  = '_($* #,##0.00_)'
        sp = row_data.get("SP500", None)
        if sp and not pd.isna(sp):
            ws.cell(row=ri, column=4, value=round(sp,2)).number_format = '_($* #,##0.00_)'
        if ri % 2 == 0:
            for c in range(1,5):
                ws.cell(row=ri, column=c).fill = PatternFill("solid", fgColor=GREY_ROW)

    # Chart
    chart = LineChart()
    chart.title          = "Portfolio vs S&P 500 vs Contributions"
    chart.y_axis.title   = "Value ($)"
    chart.height, chart.width, chart.style = 16, 32, 2
    max_r = data_start + len(sample)
    for ci, label in [(2,"Portfolio"),(3,"Contributions"),(4,"S&P 500")]:
        chart.add_data(Reference(ws, min_col=ci, min_row=data_start, max_row=max_r),
                       titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=data_start+1, max_row=max_r))
    _chart_chrome(chart, num_fmt='$#,##0')
    ws.add_chart(chart, "F2")

    _auto_width(ws)
    ws.freeze_panes = f"A{data_start+1}"
    return ws, data_start


# ── Monthly heatmap sheet ─────────────────────────────────────────────────────
def _build_heatmap_sheet(wb, heatmap_df):
    if heatmap_df is None or heatmap_df.empty:
        return
    ws = wb.create_sheet("Monthly_Returns_Heatmap")
    ws["A1"] = "Monthly Returns Heatmap (%)"
    ws["A1"].font = Font(bold=True, size=13, color=DARK_BLUE, name="Arial")
    ws.merge_cells("A1:N1")

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Full Yr"]
    ws.cell(row=2, column=1, value="Year")
    for ci, m in enumerate(months[:12], 2):
        c = ws.cell(row=2, column=ci, value=m)
        _hdr(c, bg=MID_BLUE)
    _hdr(ws.cell(row=2, column=1), bg=MID_BLUE)

    for ri, (year, row_data) in enumerate(heatmap_df.iterrows(), 3):
        ws.cell(row=ri, column=1, value=year).font = Font(bold=True, name="Arial", size=10)
        annual = 1.0
        for ci, month in enumerate(range(1,13), 2):
            val = row_data.get(month, None)
            if val is not None and not pd.isna(val):
                cell = ws.cell(row=ri, column=ci, value=round(val,2))
                cell.number_format = '0.00"%"'
                cell.font          = Font(name="Arial", size=9)
                cell.alignment     = Alignment(horizontal="center")
                colour = "C6EFCE" if val >= 0 else "FFC7CE"
                cell.fill = PatternFill("solid", fgColor=colour)
                annual *= (1 + val/100)

        ann_ret = (annual - 1) * 100
        c = ws.cell(row=ri, column=14, value=round(ann_ret,2))
        c.number_format = '0.00"%"'
        c.font          = Font(name="Arial", size=9, bold=True)
        c.alignment     = Alignment(horizontal="center")
        c.fill          = PatternFill("solid", fgColor=("C6EFCE" if ann_ret >= 0 else "FFC7CE"))

    for ci in range(1,15):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions["A"].width = 7


# ── Monte Carlo sheet ─────────────────────────────────────────────────────────
def _build_mc_sheet(wb, mc_sim_df, mc_summary, milestones):
    if mc_sim_df is None:
        return
    ws_mc = wb.create_sheet("Monte_Carlo_Portfolio")
    ws_mc["A1"] = "Portfolio Monte Carlo Simulation"
    ws_mc["A1"].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
    ws_mc.merge_cells("A1:F1")

    # Summary
    ws_mc["A2"], ws_mc["B2"] = "Metric", "Value"
    for cell in ws_mc[2]:
        _hdr(cell, bg=MID_BLUE)
    for i, (k, v) in enumerate(mc_summary.items(), 3):
        ws_mc.cell(row=i, column=1, value=k).font      = Font(name="Arial", size=10)
        ws_mc.cell(row=i, column=2, value=str(v)).font = Font(name="Arial", size=10, bold=True)
    summary_end = 3 + len(mc_summary)

    # Milestone table
    ms_start = summary_end + 2
    ws_mc.cell(row=ms_start, column=1, value="Milestone Projections").font = Font(
        bold=True, size=12, color=DARK_BLUE, name="Arial")
    ws_mc.merge_cells(f"A{ms_start}:F{ms_start}")
    ms_hdr_row = ms_start + 1
    for ci, lbl in enumerate(["Horizon","Bear (P5)","Low (P25)","Median (P50)","Bull (P75)","Best (P95)"],1):
        _hdr(ws_mc.cell(row=ms_hdr_row, column=ci, value=lbl), bg=MID_BLUE)
    for ri, (horizon, pcts) in enumerate(milestones.items(), ms_hdr_row+1):
        ws_mc.cell(row=ri, column=1, value=horizon).font = Font(name="Arial", size=10, bold=True)
        for ci, key in enumerate(["P5","P25","P50","P75","P95"],2):
            c = ws_mc.cell(row=ri, column=ci, value=pcts[key])
            c.number_format = '_($* #,##0.00_)'
            c.font          = Font(name="Arial", size=10)
            c.border        = _border()
        if ri % 2 == 0:
            for ci in range(1,7):
                ws_mc.cell(row=ri,column=ci).fill = PatternFill("solid", fgColor=GREY_ROW)

    # Percentile paths. `mc_sim_df` is the five-series percentile frame the
    # forecast keeps (columns p5..p95); a raw path matrix is still accepted and
    # reduced here, so an older caller keeps working.
    pct_col = 10
    pct_start = ms_hdr_row + len(milestones) + 3
    pct_labels = ["P5 (Bear)","P25 (Low)","P50 (Median)","P75 (Bull)","P95 (Best)"]
    ws_mc.cell(row=pct_start, column=pct_col, value="Day")
    for j, lbl in enumerate(pct_labels):
        _hdr(ws_mc.cell(row=pct_start, column=pct_col+j+1, value=lbl), bg=MID_BLUE)
    _is_pcts = list(getattr(mc_sim_df, "columns", [])) == ["p5","p25","p50","p75","p95"]
    for day_idx in range(0, len(mc_sim_df), 5):  # sample every 5 days
        _row = mc_sim_df.iloc[day_idx].values
        _vals = _row if _is_pcts else np.percentile(_row, [5,25,50,75,95])
        r = pct_start + 1 + day_idx // 5
        ws_mc.cell(row=r, column=pct_col, value=day_idx)
        for j in range(5):
            ws_mc.cell(row=r, column=pct_col+j+1,
                       value=round(float(_vals[j]), 2)).number_format = '_($* #,##0.00_)'

    # Chart
    n_rows   = len(mc_sim_df) // 5 + 1
    chart_mc = LineChart()
    chart_mc.title = "Portfolio Monte Carlo — Percentile Forecast"
    chart_mc.y_axis.title = "Portfolio Value ($)"
    chart_mc.height, chart_mc.width, chart_mc.style = 16, 32, 10
    for j in range(5):
        chart_mc.add_data(Reference(ws_mc, min_col=pct_col+j+1,
                                    min_row=pct_start, max_row=pct_start+n_rows),
                          titles_from_data=True)
    chart_mc.set_categories(Reference(ws_mc, min_col=pct_col,
                                       min_row=pct_start+1, max_row=pct_start+n_rows))
    _chart_chrome(chart_mc, num_fmt='$#,##0')
    ws_mc.add_chart(chart_mc, "A" + str(pct_start + 2))
    ws_mc.freeze_panes = f"A{ms_hdr_row+1}"
    _auto_width(ws_mc)


# ── Correlation sheet ─────────────────────────────────────────────────────────
def _build_correlation_sheet(wb, corr_matrix):
    if corr_matrix is None or corr_matrix.empty:
        return
    ws   = wb.create_sheet("Correlation_Matrix")
    labs = list(corr_matrix.columns)
    ws.cell(row=1,column=1,value="Correlation Matrix (Daily Returns)").font = Font(
        bold=True, size=12, color=DARK_BLUE, name="Arial")
    ws.merge_cells(f"A1:{get_column_letter(len(labs)+1)}1")
    for ci, lbl in enumerate(labs,2):
        _hdr(ws.cell(row=2,column=ci,value=lbl), bg=MID_BLUE)
    for ri, lbl in enumerate(labs,3):
        _hdr(ws.cell(row=ri,column=1,value=lbl), bg=MID_BLUE)
        for ci, col_lbl in enumerate(labs,2):
            val  = corr_matrix.loc[lbl,col_lbl]
            cell = ws.cell(row=ri,column=ci,value=round(float(val),4))
            cell.number_format = "0.0000"
            cell.font          = Font(name="Arial", size=10)
            cell.border        = _border()
            cell.alignment     = Alignment(horizontal="center")
    data_range = f"B3:{get_column_letter(len(labs)+1)}{len(labs)+2}"
    ws.conditional_formatting.add(data_range, ColorScaleRule(
        start_type="num",start_value=-1,start_color="FF9999",
        mid_type="num",mid_value=0,mid_color="FFFFFF",
        end_type="num",end_value=1,end_color="99CCFF"))
    _auto_width(ws)


# ── Master builder ────────────────────────────────────────────────────────────
def build_portfolio_excel(preferences, final_weights, stock_metrics,
                           backtest_df, backtest_metrics, heatmap_df,
                           mc_sim_df, mc_summary, milestones,
                           corr_matrix, diversification_score,
                           ticker_info=None):
    if ticker_info is None:
        ticker_info = {}

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, preferences, final_weights, stock_metrics,
                     backtest_metrics, mc_summary, diversification_score)
    _build_holdings_sheet(wb, final_weights, stock_metrics, ticker_info)
    _build_backtest_sheet(wb, backtest_df, backtest_metrics)
    _build_heatmap_sheet(wb, heatmap_df)
    _build_mc_sheet(wb, mc_sim_df, mc_summary, milestones)
    _build_correlation_sheet(wb, corr_matrix)

    sheets_so_far = list(wb.sheetnames)
    _build_cover(wb, preferences, final_weights, backtest_metrics, mc_summary, sheets_so_far)

    desired = ["Cover","Dashboard","Holdings_Breakdown","Backtest_Results",
               "Monthly_Returns_Heatmap","Monte_Carlo_Portfolio","Correlation_Matrix"]
    existing = wb.sheetnames
    ordered  = [s for s in desired if s in existing]
    extras   = [s for s in existing if s not in ordered]
    for i, name in enumerate(ordered + extras):
        wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# Tracked-portfolio report — the "Your Portfolios" page export.
# Input is the (already cached) tracker.track_portfolio result plus per-ticker
# profiles from market_data.get_ticker_profiles. Every profile field may be None;
# each sheet renders an em-dash rather than assuming coverage.
# ═══════════════════════════════════════════════════════════════════════════════

DASH = "—"


def _num(ws, row, col, val, fmt=None, bold=False, fill=None):
    """Numeric cell, or a right-aligned grey em-dash when the value is missing."""
    if val is None or (isinstance(val, float) and val != val):
        c = ws.cell(row=row, column=col, value=DASH)
        c.font = Font(name="Arial", size=10, color="999999")
    else:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", size=10, bold=bold)
        if fmt:
            c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    c.border = _border()
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


def _tp_title(ws, title, sub_lines, ncols=6):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = Font(size=16, bold=True,
                                                      color=DARK_BLUE, name="Arial")
    ws.row_dimensions[1].height = 26
    for i, line in enumerate(sub_lines, 2):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ncols)
        ws.cell(row=i, column=1, value=line).font = Font(size=9, italic=True,
                                                         color="888888", name="Arial")
    return 2 + len(sub_lines) + 1   # first free row


def _finite(v):
    """Numbers only — NaN and inf become None so callers render a dash."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if f == f and abs(f) != float("inf") else None


def _derive_portfolio_stats(tracked, profiles):
    """One shared analytics pass every sheet reads from.

    Weighted P/E is harmonic (earnings-weighted) over positive-P/E holdings —
    an arithmetic mean of P/Es lets one richly valued position dominate.
    Dividend yield counts non-payers as zero (a portfolio's yield includes
    them); growth is renormalised over covered weight because a missing growth
    figure is unknown, not zero.
    """
    from portfolio_analysis import portfolio_beta

    rows = []
    for h in tracked["holdings"]:
        p = profiles.get(h["ticker"], {}) or {}
        rows.append({**h, "w": h["weight_pct"] / 100.0,
                     "name":       p.get("name") or h["ticker"],
                     "sector":     p.get("sector") or "Unknown",
                     "industry":   p.get("industry") or "Unknown",
                     "pe":         p.get("pe"),
                     "div_yield":  p.get("div_yield"),
                     "beta":       p.get("beta"),
                     "rev_growth": p.get("rev_growth"),
                     "eps_growth": p.get("eps_growth"),
                     # Carried so callers can weight by company size (the
                     # large/mid/small profile) without re-reading `profiles`.
                     "market_cap": p.get("market_cap")})
    rows.sort(key=lambda r: -r["w"])

    total_value = sum(r["value"] for r in rows)
    weights = {r["ticker"]: r["w"] for r in rows}

    def _wavg(field):
        cov = [(r["w"], r[field]) for r in rows if r[field] is not None]
        wsum = sum(w for w, _ in cov)
        return (sum(w * v for w, v in cov) / wsum if wsum > 0 else None), wsum

    betas = {r["ticker"]: r["beta"] for r in rows if r["beta"] is not None}
    beta_cov = sum(r["w"] for r in rows if r["beta"] is not None)
    port_beta = portfolio_beta(weights, betas) if betas else None

    pe_rows = [(r["w"], r["pe"]) for r in rows if r["pe"] and r["pe"] > 0]
    pe_cov = sum(w for w, _ in pe_rows)
    wpe = pe_cov / sum(w / pe for w, pe in pe_rows) if pe_cov > 0 else None

    div_yield = sum(r["w"] * (r["div_yield"] or 0.0) for r in rows)
    eps_g, eps_cov = _wavg("eps_growth")
    rev_g, rev_cov = _wavg("rev_growth")

    def _group(field):
        d = {}
        for r in rows:
            d[r[field]] = d.get(r[field], 0.0) + r["value"]
        return sorted(d.items(), key=lambda kv: -kv[1])

    sectors, industries = _group("sector"), _group("industry")
    ws_ = [r["w"] for r in rows]
    hhi = sum(w * w for w in ws_)

    return {
        "rows": rows, "total_value": total_value, "n": len(rows),
        "largest": rows[0] if rows else None,
        "top5":  sum(ws_[:5]),
        "top10": sum(ws_[:10]),
        "beta": port_beta, "beta_cov": beta_cov,
        "wpe": wpe, "pe_cov": pe_cov,
        "div_yield": div_yield,
        "eps_growth": eps_g, "eps_cov": eps_cov,
        "rev_growth": rev_g, "rev_cov": rev_cov,
        "sectors": sectors, "industries": industries,
        "hhi": hhi, "eff_n": (1.0 / hhi if hhi > 0 else None),
        # NaN, not None, is what a volatility computed from two observations
        # comes back as — and it printed straight through to the report as
        # "nan%". Anything non-finite is treated as absent.
        "vol": _finite(tracked["metrics"].get("Ann. Volatility")),
        "n_days": len(tracked["curve"]),
        # Below this, ratios and drawdowns describe one market episode rather
        # than the portfolio. Reports use it to suppress statistics that would
        # otherwise read as fact ("deepest decline 0.0%", "100% positive months").
        "thin_history": len(tracked["curve"]) < 21,
        # Did company metadata actually resolve? Without this the reports cannot
        # tell "no sector" from "the sector lookup failed".
        "profiles_ok": sum(
            1 for r in rows if r.get("sector") not in (None, "Unknown")),
    }


def _risk_observations(s, metrics):
    """Concise, professional commentary — statements a human analyst would sign."""
    obs = []
    n, hhi, eff_n = s["n"], s["hhi"], s["eff_n"]
    if hhi < 0.10 and n >= 15:
        obs.append(f"The portfolio is well diversified: {n} positions with an "
                   f"effective holding count of ~{eff_n:.0f} once weights are accounted for.")
    elif hhi <= 0.18:
        obs.append(f"Diversification is moderate: {n} positions, but weight is "
                   f"concentrated enough that the portfolio behaves like ~{eff_n:.0f} "
                   f"equal-sized holdings.")
    else:
        obs.append(f"The portfolio has elevated concentration risk: despite {n} "
                   f"position{'s' if n != 1 else ''}, it behaves like ~{max(eff_n or 1, 1):.0f} "
                   f"equal-sized holdings.")
    lg = s["largest"]
    if lg:
        line = (f"The largest position, {lg['ticker']}, represents {lg['w']:.1%} of assets"
                f"{' — outsized single-name exposure.' if lg['w'] > 0.25 else '.'}")
        obs.append(line)
    if s["top5"] > 0.60 and n > 5:
        obs.append(f"The top 5 holdings account for {s['top5']:.1%} of the portfolio — "
                   f"performance will be driven largely by these names.")
    for sec, val in s["sectors"]:
        w = val / s["total_value"] if s["total_value"] else 0
        if w > 0.30 and sec not in ("Unknown", "Fund / ETF"):
            obs.append(f"{sec} exposure is {w:.1%} of assets, exceeding the 30% level "
                       f"typically treated as a sector concentration.")
    fund_w = sum(val for sec, val in s["sectors"] if sec == "Fund / ETF")
    if s["total_value"] and fund_w / s["total_value"] > 0.40:
        obs.append(f"Funds/ETFs make up {fund_w / s['total_value']:.1%} of assets; "
                   f"look-through sector exposure is broader than the single-name "
                   f"figures above suggest.")
    if s["beta"] is not None:
        if s["beta"] > 1.2:
            obs.append(f"A weighted beta of {s['beta']:.2f} implies meaningfully more "
                       f"market sensitivity than the S&P 500.")
        elif s["beta"] < 0.8:
            obs.append(f"A weighted beta of {s['beta']:.2f} gives the portfolio a "
                       f"defensive tilt relative to the S&P 500.")
    if s["n_days"] < 63:
        obs.append(f"Volatility and risk statistics are measured over only "
                   f"{s['n_days']} trading days since inception and should be read "
                   f"as indicative, not established.")
    if s["beta_cov"] < 0.9 and s["beta"] is not None:
        obs.append("Positions without a published beta are assumed to have β = 1.0 "
                   "in the portfolio figure.")
    return obs


def _intel_insights(s):
    """(label, commentary) pairs — what the portfolio actually looks like."""
    rows, out = s["rows"], []
    lg = s["largest"]
    if lg:
        out.append(("Position Intelligence", None))
        out.append(("Largest holding",
                    f"{lg['ticker']} ({lg['name']}) at {lg['w']:.1%} of the portfolio "
                    f"(${lg['value']:,.0f})."))
        gains = [r for r in rows if r.get("gain_pct") is not None]
        if gains:
            best = max(gains, key=lambda r: r["gain_pct"])
            worst = min(gains, key=lambda r: r["gain_pct"])
            out.append(("Best performer since inception",
                        f"{best['ticker']} is {'up' if best['gain_pct'] >= 0 else 'down'} "
                        f"{abs(best['gain_pct']):.1f}% since it was added."))
            if worst is not best:
                out.append(("Laggard",
                            f"{worst['ticker']} is {'down' if worst['gain_pct'] < 0 else 'up only'} "
                            f"{abs(worst['gain_pct']):.1f}% since it was added."))

    out.append(("Fundamental Standouts", None))
    def _pick(field, top=True):
        c = [r for r in rows if r[field] is not None]
        return (max if top else min)(c, key=lambda r: r[field]) if c else None
    fg = _pick("rev_growth")
    if fg and fg["rev_growth"] > 0:
        out.append(("Fastest-growing company",
                    f"{fg['ticker']} — revenue up {fg['rev_growth']:.1%} year over year."))
    dp = _pick("div_yield")
    if dp and (dp["div_yield"] or 0) > 0.001:
        out.append(("Highest dividend payer",
                    f"{dp['ticker']} yields {dp['div_yield']:.2%}, against a portfolio "
                    f"yield of {s['div_yield']:.2%}."))
    hb, db = _pick("beta"), _pick("beta", top=False)
    if hb:
        out.append(("Highest-beta holding",
                    f"{hb['ticker']} (β {hb['beta']:.2f}) contributes the most market "
                    f"sensitivity."))
    if db and hb is not db:
        out.append(("Most defensive holding",
                    f"{db['ticker']} (β {db['beta']:.2f}) is the steadiest name in the book."))

    out.append(("Composition", None))
    if s["sectors"] and s["total_value"]:
        sec, val = s["sectors"][0]
        out.append(("Largest sector",
                    f"{sec} at {val / s['total_value']:.1%} of assets"
                    + (f", roughly {val / s['total_value'] * s['n']:.1f}× an equal spread "
                       f"across this portfolio's holdings." if s["n"] > 1 else ".")))
    if s["eff_n"] is not None:
        out.append(("Diversification",
                    f"{s['n']} positions behaving like ~{s['eff_n']:.0f} equal-sized "
                    f"holdings (HHI {s['hhi']:.3f})."))
    if s["wpe"] is not None:
        out.append(("Valuation profile",
                    f"Weighted P/E of {s['wpe']:.1f}× across the {s['pe_cov']:.0%} of the "
                    f"portfolio with positive earnings."))
    return out


# ── Sheet 1 — Portfolio Summary ───────────────────────────────────────────────
def _build_tp_summary(wb, name, tracked, s):
    ws = wb.create_sheet("Portfolio_Summary")
    m = tracked["metrics"]
    row = _tp_title(ws, "◈ QuantWizard — Portfolio Report",
                    [f"{name}  ·  tracked since {tracked['inception_date']}",
                     f"Generated {datetime.now().strftime('%B %d, %Y')}  |  "
                     f"Data: Yahoo Finance · Polygon  |  Not investment advice"])
    for col, w in zip("ABCDEFGH", [30, 16, 3, 14, 20, 13, 4, 12]):
        ws.column_dimensions[col].width = w

    _sec_hdr(ws, row, "Portfolio Snapshot", col_end=2); row += 1
    lg = s["largest"]
    pairs = [
        ("Total Portfolio Value", s["total_value"], '_($* #,##0_)'),
        ("Number of Holdings",    s["n"], '0'),
        ("Largest Position",      f"{lg['ticker']}  ({lg['w']:.1%})" if lg else DASH, None),
        ("Top 5 Concentration",   s["top5"], '0.0%'),
        ("Portfolio Beta",        s["beta"], '0.00'),
        ("Ann. Volatility",       (s["vol"] / 100 if s["vol"] is not None else None), '0.0%'),
        ("Dividend Yield",        s["div_yield"], '0.00%'),
        ("Weighted P/E",          s["wpe"], '0.0"×"'),
        ("Weighted Earnings Growth (YoY)", s["eps_growth"], '0.0%'),
        ("Sectors Represented",   len([1 for k, _ in s["sectors"] if k != "Unknown"]), '0'),
        ("Industries Represented", len([1 for k, _ in s["industries"] if k != "Unknown"]), '0'),
        ("Total Return (since inception)",
         (m.get("Total Return", 0) or 0) / 100, '+0.0%;-0.0%'),
        ("vs S&P 500 (same money)",
         (m["vs S&P 500"] / 100 if isinstance(m.get("vs S&P 500"), (int, float)) else None),
         '+0.0%;-0.0%'),
    ]
    for label, val, fmt in pairs:
        if isinstance(val, str):
            _kv(ws, row, label, val)
        else:
            ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10)
            ws.cell(row=row, column=1).border = _border()
            _num(ws, row, 2, val, fmt, bold=True)
            if row % 2 == 0:
                for c in (1, 2):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=GREY_ROW)
        row += 1

    # Top holdings table — doubles as the pie chart's data range.
    th_hdr = row + 1
    _sec_hdr(ws, th_hdr, "Top Holdings", col_start=4, col_end=6)
    for ci, h in enumerate(["Ticker", "Company", "Value"], 4):
        _hdr(ws.cell(row=th_hdr + 1, column=ci, value=h), bg=MID_BLUE)
    top = s["rows"][:8]
    r = th_hdr + 2
    for h in top:
        ws.cell(row=r, column=4, value=h["ticker"]).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=5, value=h["name"][:28]).font = Font(name="Arial", size=10)
        _num(ws, r, 6, h["value"], '_($* #,##0_)')
        for c in (4, 5):
            ws.cell(row=r, column=c).border = _border()
        if r % 2 == 0:
            for c in (4, 5, 6):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREY_ROW)
        r += 1
    rest = s["rows"][8:]
    if rest:
        ws.cell(row=r, column=4, value=f"Other ({len(rest)})").font = Font(
            name="Arial", size=10, italic=True)
        ws.cell(row=r, column=5, value="").border = _border()
        ws.cell(row=r, column=4).border = _border()
        _num(ws, r, 6, sum(h["value"] for h in rest), '_($* #,##0_)')
        r += 1

    pie = PieChart()
    pie.title = "Asset Allocation"
    pie.height, pie.width = 9, 13
    pie.add_data(Reference(ws, min_col=6, min_row=th_hdr + 1, max_row=r - 1),
                 titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=4, min_row=th_hdr + 2, max_row=r - 1))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.roundedCorners = False
    pie.graphical_properties = GraphicalProperties()
    pie.graphical_properties.line.noFill = True
    if pie.legend is not None:
        pie.legend.position = "r"
        pie.legend.overlay = False
    ws.add_chart(pie, f"D{r + 2}")
    ws.freeze_panes = "A4"


# ── Sheet 2 — Holdings Analysis ───────────────────────────────────────────────
def _build_tp_holdings(wb, s):
    ws = wb.create_sheet("Holdings_Analysis")
    headers = ["Ticker", "Company", "Shares", "Price", "Market Value", "Weight",
               "Sector", "Industry", "P/E", "Rev Growth", "EPS Growth",
               "Div Yield", "Beta"]
    ws.append(headers)
    _style_header_row(ws, bg=DARK_BLUE)

    fmts = {3: '#,##0.00', 4: '_($* #,##0.00_)', 5: '_($* #,##0_)', 6: '0.0%',
            9: '0.0', 10: '0.0%', 11: '0.0%', 12: '0.00%', 13: '0.00'}
    for ri, h in enumerate(s["rows"], 2):
        vals = [h["ticker"], h["name"][:34], h["shares"], h["last_price"], h["value"],
                h["w"], h["sector"], h["industry"], h["pe"], h["rev_growth"],
                h["eps_growth"], h["div_yield"], h["beta"]]
        for ci, v in enumerate(vals, 1):
            if ci in fmts:
                _num(ws, ri, ci, v, fmts[ci])
            else:
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = Font(name="Arial", size=10, bold=(ci == 1))
                c.border = _border()
        if ri % 2 == 0:
            for ci in range(1, 14):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=GREY_ROW)

    last = ws.max_row
    ws.conditional_formatting.add(
        f"F2:F{last}", DataBarRule(start_type="num", start_value=0,
                                   end_type="max", color=MID_BLUE))
    for col in ("J", "K"):   # growth columns: red below zero, green above
        ws.conditional_formatting.add(
            f"{col}2:{col}{last}",
            ColorScaleRule(start_type="min", start_color="F4CCCC",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="max", end_color="D9EAD3"))
    _auto_width(ws, max_w=34)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:M{last}"


# ── Sheet 3 — Allocation Analysis ─────────────────────────────────────────────
def _build_tp_allocation(wb, s):
    ws = wb.create_sheet("Allocation_Analysis")
    row = _tp_title(ws, "Allocation Analysis",
                    ["By security, sector and industry — dollar value and weight"], ncols=4)
    ws.column_dimensions["A"].width = 30
    for col in "BC":
        ws.column_dimensions[col].width = 15

    total = s["total_value"] or 1.0
    # Each row carries a long label for the TABLE and a short one for the CHART.
    # The charts previously took their categories from column A, so every bar was
    # labelled "SPY — State Street SPDR S&P   ◂ largest" — the sort marker and
    # the company name both leaked onto the axis and squeezed the plot.
    blocks = [
        ("By Security", [(f"{h['ticker']} — {h['name'][:22]}", h["value"], h["ticker"])
                         for h in s["rows"]]),
        ("By Sector",   [(k, v, k) for k, v in s["sectors"]]),
        ("By Industry", [(k, v, k[:22]) for k, v in s["industries"]]),
        # (third element is unused now that charts read column A, but kept so the
        # unpacking below stays a single shape across all three blocks)
    ]
    anchors = {}
    for title, items in blocks:
        _sec_hdr(ws, row, title, col_end=3); row += 1
        for ci, h in enumerate(["Name", "Value", "Weight"], 1):
            _hdr(ws.cell(row=row, column=ci, value=h), bg=MID_BLUE)
        hdr = row; row += 1
        for i, (label, val, short) in enumerate(items):
            top = (i == 0)   # lists arrive sorted desc — the largest gets flagged
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(name="Arial", size=10, bold=top)
            c.border = _border()
            _num(ws, row, 2, val, '_($* #,##0_)', bold=top)
            _num(ws, row, 3, val / total, '0.0%', bold=top)
            if top:
                for ci in (1, 2, 3):
                    ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=LIGHT_BG)
            elif row % 2 == 0:
                for ci in (1, 2, 3):
                    ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=GREY_ROW)
            row += 1
        anchors[title] = (hdr, row - 1)
        row += 2

    sec_hdr, sec_last = anchors["By Sector"]
    ch = BarChart(); ch.type = "col"
    ch.title = "Sector Allocation"
    ch.height, ch.width, ch.legend = 8, 15, None
    ch.add_data(Reference(ws, min_col=2, min_row=sec_hdr, max_row=sec_last),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=sec_hdr + 1, max_row=sec_last))
    # Column 2 is the dollar value, so the scale is money, not a percentage.
    _chart_chrome(ch, bar_colour="3F6C9C", num_fmt='$#,##0')
    ws.add_chart(ch, "E4")

    pos_hdr, pos_last = anchors["By Security"]
    ch2 = BarChart(); ch2.type = "bar"
    ch2.title = "Position Weights"
    ch2.height, ch2.width, ch2.legend = max(8, min(20, s["n"] * 0.55)), 15, None
    ch2.add_data(Reference(ws, min_col=3, min_row=pos_hdr, max_row=pos_last),
                 titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=pos_hdr + 1, max_row=pos_last))
    # Column 3 is the weight. Gridlines off: a horizontal bar chart already has
    # the category rules, and a second set behind them just adds noise.
    _chart_chrome(ch2, bar_colour="3F6C9C", num_fmt='0.0%', gridlines=False)
    ws.add_chart(ch2, "E22")
    ws.freeze_panes = "A4"


# ── Sheet 4 — Risk & Diversification ──────────────────────────────────────────
def _build_tp_risk(wb, tracked, s):
    ws = wb.create_sheet("Risk_Diversification")
    m = tracked["metrics"]
    row = _tp_title(ws, "Risk & Diversification",
                    [f"Statistics measured over {s['n_days']} trading days since "
                     f"{tracked['inception_date']}"], ncols=4)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    for col in "CD":
        ws.column_dimensions[col].width = 22

    _sec_hdr(ws, row, "Risk Metrics", col_end=2); row += 1
    conc_label = ("Low" if s["hhi"] < 0.10 else "Moderate" if s["hhi"] <= 0.18 else "High")
    lg = s["largest"]
    pairs = [
        ("Portfolio Beta",          s["beta"], '0.00'),
        ("Ann. Volatility",         (s["vol"] / 100 if s["vol"] is not None else None), '0.0%'),
        ("Sharpe Ratio",            m.get("Sharpe Ratio"), '0.00'),
        ("Max Drawdown",            (m.get("Max Drawdown", 0) or 0) / 100, '0.0%'),
        ("Concentration (HHI)",     s["hhi"], '0.000'),
        ("Concentration Level",     conc_label, None),
        ("Effective # of Holdings", s["eff_n"], '0.0'),
        ("Largest Position",        lg["w"] if lg else None, '0.0%'),
        ("Top 5 Concentration",     s["top5"], '0.0%'),
        ("Top 10 Concentration",    s["top10"], '0.0%'),
    ]
    for label, val, fmt in pairs:
        if isinstance(val, str):
            _kv(ws, row, label, val)
        else:
            ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10)
            ws.cell(row=row, column=1).border = _border()
            _num(ws, row, 2, val, fmt, bold=True)
            if row % 2 == 0:
                for c in (1, 2):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=GREY_ROW)
        row += 1

    row += 1
    _sec_hdr(ws, row, "Observations", col_end=4); row += 1
    for i, text in enumerate(_risk_observations(s, m)):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"•  {text}")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 1:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=GREY_ROW)
        ws.row_dimensions[row].height = max(15, 13 * (len(text) // 80 + 1))
        row += 1
    ws.freeze_panes = "A4"


# ── Sheet 5 — Portfolio Intelligence ──────────────────────────────────────────
def _build_tp_intel(wb, s):
    ws = wb.create_sheet("Portfolio_Intelligence")
    row = _tp_title(ws, "Portfolio Intelligence",
                    ["What the portfolio actually looks like — generated from your "
                     "holdings, not a template"], ncols=5)
    ws.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws.column_dimensions[col].width = 18

    for label, text in _intel_insights(s):
        if text is None:                      # section header row
            _sec_hdr(ws, row, label, col_end=5)
            row += 1
            continue
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(vertical="top")
        c.border = _border()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        t = ws.cell(row=row, column=2, value=text)
        t.font = Font(name="Arial", size=10)
        t.alignment = Alignment(wrap_text=True, vertical="top")
        t.border = _border()
        ws.row_dimensions[row].height = max(15, 13 * (len(text) // 70 + 1))
        if row % 2 == 0:
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=GREY_ROW)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1,
            value="Fundamental data: Yahoo Finance, as of report generation. Figures "
                  "may be delayed or unavailable for some securities. For information "
                  "only — not investment advice.").font = Font(name="Arial", size=8,
                                                               italic=True, color="888888")


# ── Master builder ────────────────────────────────────────────────────────────
def _build_tp_live(wb, portfolio_name, s):
    """A self-refreshing valuation sheet — prices come from Excel, not from us.

    Every other sheet is a snapshot taken when the file was generated. This one
    re-prices the same holdings each time it recalculates, using Excel's own
    STOCKHISTORY function, so the workbook stays useful the day after it was
    downloaded instead of going stale immediately.

    Why this and not a web query: a Power Query connection lives in a binary
    DataMashup part that openpyxl cannot write, and a WEBSERVICE() call is
    blocked in Excel for the web and in most corporate builds — and would put
    an API key inside a file the user then emails around. STOCKHISTORY needs no
    backend, no key, and no service of ours to stay up. It is a Microsoft 365
    function, so the sheet degrades to a clear message elsewhere rather than a
    wall of #NAME? errors.

    The shares are hardcoded because they are facts about the position; only the
    price, and everything derived from it, is live.
    """
    ws = wb.create_sheet("Live_Prices")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 12
    for col, wdt in zip("BCDEFG", (13, 15, 15, 15, 15, 12)):
        ws.column_dimensions[col].width = wdt

    ws["A1"] = f"{portfolio_name} — Live Valuation"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=DARK_BLUE)
    ws["A2"] = ("Prices refresh from Excel itself. Press F9, or Data ▸ Refresh All, "
                "to re-price the portfolio — no need to download the report again.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws["A3"] = ("Requires Microsoft 365 (STOCKHISTORY). In older Excel the price "
                "column will read “Needs Microsoft 365”; every other sheet still works.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="808080")

    ws["A5"] = "Last recalculated"
    ws["A5"].font = Font(name="Arial", size=9, color="808080")
    ws["B5"] = "=TEXT(NOW(),\"yyyy-mm-dd hh:mm\")"
    ws["B5"].font = Font(name="Arial", size=9, bold=True)

    hdr_row = 7
    headers = ["Ticker", "Shares", "Cost basis", "Live price", "Live value",
               "Gain / loss", "Return"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws.cell(row=hdr_row, column=ci, value=h))

    first = hdr_row + 1
    rows = s["rows"]
    for i, h in enumerate(rows):
        r = first + i
        ws.cell(row=r, column=1, value=h["ticker"]).font = Font(name="Arial", size=10, bold=True)
        _num(ws, r, 2, h["shares"], "#,##0.000")
        _num(ws, r, 3, h.get("cost_basis"), '_($* #,##0_)')
        # LET keeps STOCKHISTORY to a single call; INDEX takes the last row of
        # the returned array, which is the most recent close it has.
        ws.cell(row=r, column=4, value=(
            f'=IFERROR(LET(h,STOCKHISTORY(A{r},TODAY()-10,TODAY(),0,0,1),'
            f'INDEX(h,ROWS(h),1)),"Needs Microsoft 365")'))
        ws.cell(row=r, column=4).number_format = '_($* #,##0.00_)'
        ws.cell(row=r, column=5, value=f"=IFERROR(B{r}*D{r},\"\")")
        ws.cell(row=r, column=5).number_format = '_($* #,##0_)'
        ws.cell(row=r, column=6, value=f"=IFERROR(E{r}-C{r},\"\")")
        ws.cell(row=r, column=6).number_format = '_($* #,##0_)'
        ws.cell(row=r, column=7, value=f"=IFERROR(E{r}/C{r}-1,\"\")")
        ws.cell(row=r, column=7).number_format = '+0.0%;-0.0%'
        for ci in range(1, 8):
            c = ws.cell(row=r, column=ci)
            c.border = _border()
            if ci >= 4:
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="right")
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=GREY_ROW)

    last = first + len(rows) - 1
    tot = last + 1
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
    # Guarded on COUNT: where STOCKHISTORY is unavailable the per-row values are
    # empty text, so a plain SUM totalled to $0 against a real cost basis and the
    # return read -100.0% — a portfolio that looks wiped out rather than a
    # feature that isn't available.
    _live = f"COUNT(E{first}:E{last})=0"
    for ci, formula in ((3, f"=SUM(C{first}:C{last})"),
                        (5, f'=IF({_live},"",SUM(E{first}:E{last}))'),
                        (6, f'=IF({_live},"",SUM(F{first}:F{last}))')):
        c = ws.cell(row=tot, column=ci, value=formula)
        c.number_format = '_($* #,##0_)'
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=tot, column=7,
                value=f'=IF({_live},"",IFERROR(E{tot}/C{tot}-1,""))')
    c.number_format = '+0.0%;-0.0%'
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="right")
    for ci in range(1, 8):
        cc = ws.cell(row=tot, column=ci)
        cc.border = _border()
        cc.fill = PatternFill("solid", fgColor="D6E4F0")

    note = tot + 2
    ws.cell(row=note, column=1, value=(
        "Cost basis is what was paid for the shares still held, so gain / loss "
        "here is unrealised. Prices are end-of-day from Microsoft's market data "
        "provider and may lag intraday."))
    ws.cell(row=note, column=1).font = Font(name="Arial", size=9, italic=True,
                                            color="808080")
    # The path that works for everyone, M365 or not.
    lc = ws.cell(row=note + 2, column=1,
                 value="↗  Open this portfolio on QuantWizard for a fully "
                       "refreshed report")
    lc.font = Font(name="Arial", size=10, bold=True, color=MID_BLUE,
                   underline="single")
    lc.hyperlink = "https://quantwizard.co/?page=portfolios"
    ws.freeze_panes = f"A{first}"
    return ws


def build_tracked_portfolio_excel(portfolio_name, tracked, profiles):
    """The Your Portfolios export: tracker.track_portfolio result + profiles → xlsx buffer."""
    s = _derive_portfolio_stats(tracked, profiles)
    wb = Workbook()
    wb.remove(wb.active)
    _build_tp_summary(wb, portfolio_name, tracked, s)
    _build_tp_holdings(wb, s)
    _build_tp_live(wb, portfolio_name, s)
    _build_tp_allocation(wb, s)
    _build_tp_risk(wb, tracked, s)
    _build_tp_intel(wb, s)
    for name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = DARK_BLUE
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

import io
import os
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                   ParagraphProperties, CharacterProperties)

# Hairline grey for gridlines and axis rules. Excel's default is near-black,
# which competes with the data it is supposed to sit behind.
GRID_GREY = "D9D9D9"
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from constants import get_risk_free_rate
from disclaimers import SHORT as DISCLAIMER_SHORT
from market_data import consensus_from_recommendation
from analysis import compute_scorecard

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    MPL_AVAILABLE = True
except ImportError:
    MPL_AVAILABLE = False

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
GREEN_OK   = "70AD47"
RED_BAD    = "FF0000"
WHITE      = "FFFFFF"
GREY_ROW   = "F2F2F2"
TILE_BG    = "F0F5FB"   # pale blue KPI-tile fill
BAD_FILL   = "FFC7CE"   # Excel-classic light red
BAD_TEXT   = "9C0006"   # Excel-classic dark red
GOOD_FILL  = "C6EFCE"   # Excel-classic light green
GOOD_TEXT  = "006100"   # Excel-classic dark green
INPUT_BG   = "FFF2CC"   # editable-assumption cells on the Valuation model

# Custom number formats used by the live valuation model. Values stay numeric
# (so formulas can reference them) while displaying in the same $B / % style as
# the static cells elsewhere in the workbook.
FMT_USD    = '_($* #,##0.00_)'
FMT_BN     = '$#,##0.0,,,"B"'
FMT_SHARES = '#,##0.00,,,"B"'
FMT_PCT1   = '0.0%'
FMT_PCT2   = '0.00%'
FMT_SIGNED = '+0.0%;-0.0%'

# Kept in sync with the PowerPoint deck's data-source line (pptx_builder.py).
DATA_SOURCE_LINE = "Polygon · Yahoo Finance · Finnhub · SEC EDGAR"

# Where a reader goes to get today's version of this workbook. Kept as one
# constant so a domain change is a single edit rather than a hunt through
# generated strings.
LIVE_BASE = "https://quantwizard.co"

# Full logo (light line-art on transparent) — shows on the navy cover band.
_ASSET_LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo_full.png")


def _border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _hdr_cell(cell, bg=DARK_BLUE, fg=WHITE):
    cell.font      = Font(bold=True, color=fg, name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


def style_header_row(ws, bg=DARK_BLUE):
    for cell in ws[1]:
        _hdr_cell(cell, bg=bg)


def auto_col_width(ws, max_w=28):
    from openpyxl.cell.cell import Cell
    for col in ws.columns:
        real_cells = [c for c in col if isinstance(c, Cell)]
        if not real_cells:
            continue
        best = max((len(str(c.value or "")) for c in real_cells), default=10)
        ws.column_dimensions[real_cells[0].column_letter].width = min(best + 3, max_w)


def make_sparkline(values, color="#2E75B6", width=2.2, height=0.45):
    if not MPL_AVAILABLE:
        return None
    vals = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
    if len(vals) < 2:
        return None
    fig, ax = plt.subplots(figsize=(width, height))
    ax.plot(vals, color=color, linewidth=1.2)
    ax.fill_between(range(len(vals)), vals, min(vals), alpha=0.15, color=color)
    ax.set_axis_off()
    fig.patch.set_alpha(0)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=72, bbox_inches="tight", transparent=True, pad_inches=0)
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Cover page ────────────────────────────────────────────────────────────────
def _build_cover(wb, ticker, period, sheetnames, df=None):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 21

    # Letterhead — the logo on a navy band directly above the title, so the two
    # navy bands read as one continuous branded header.
    ws.merge_cells("B1:E1")
    ws["B1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws.row_dimensions[1].height = 54
    if os.path.exists(_ASSET_LOGO):
        try:
            _logo = XLImage(_ASSET_LOGO)
            _asp  = (_logo.width / _logo.height) if _logo.height else 1.47
            _logo.height = 48
            _logo.width  = int(48 * _asp)
            ws.add_image(_logo, "B1")
        except Exception:
            pass

    # Title band — solid navy with white text, matching the deck cover.
    ws.merge_cells("B2:E3")
    c = ws["B2"]
    c.value     = f"{ticker}  —  Equity Research Report"
    c.font      = Font(size=22, bold=True, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 14

    ws.merge_cells("B4:E4")
    c = ws["B4"]
    c.value     = (f"Period: {period}    |    Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
                   f"    |    Multi-source data: {DATA_SOURCE_LINE}")
    c.font      = Font(size=9, italic=True, color="888888", name="Calibri")
    c.alignment = Alignment(horizontal="left", indent=1)

    # KPI band — four headline stats so the cover reads like a tear-sheet, not a TOC.
    toc_start = 6
    if df is not None and len(df) > 1:
        latest, firstrow = df.iloc[-1], df.iloc[0]
        period_ret = latest["Close"] / firstrow["Close"] - 1
        ret     = df["Daily_Return"].dropna()
        ann_vol = ret.std() * np.sqrt(252)
        ann_ret = ret.mean() * 252
        sharpe  = (ann_ret - get_risk_free_rate()) / ann_vol if ann_vol else float("nan")
        tiles = [
            ("Current Price",   f"${latest['Close']:,.2f}",                    DARK_BLUE),
            ("Period Return",   f"{period_ret * 100:+.1f}%",
                                GREEN_OK if period_ret >= 0 else RED_BAD),
            ("Sharpe Ratio",    f"{sharpe:.2f}" if pd.notna(sharpe) else "N/A",
                                GREEN_OK if pd.notna(sharpe) and sharpe >= 1 else DARK_BLUE),
            ("Ann. Volatility", f"{ann_vol * 100:.1f}%",                       DARK_BLUE),
        ]
        for i, (label, value, accent) in enumerate(tiles):
            col   = 2 + i   # B, C, D, E
            vcell = ws.cell(row=6, column=col, value=value)
            vcell.font      = Font(size=16, bold=True, name="Calibri", color=accent)
            vcell.alignment = Alignment(horizontal="center", vertical="center")
            vcell.fill      = PatternFill("solid", fgColor=TILE_BG)
            vcell.border    = _border()
            lcell = ws.cell(row=7, column=col, value=label.upper())
            lcell.font      = Font(size=8, bold=True, name="Calibri", color="808080")
            lcell.alignment = Alignment(horizontal="center", vertical="center")
            lcell.fill      = PatternFill("solid", fgColor=TILE_BG)
            lcell.border    = _border()
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 16
        toc_start = 9

    ws.cell(row=toc_start, column=2, value="TABLE OF CONTENTS").font = \
        Font(bold=True, size=12, color=DARK_BLUE, name="Calibri")
    for i, name in enumerate(sheetnames, toc_start + 1):
        cell = ws.cell(row=i, column=2, value=name.replace("_", " "))
        cell.font      = Font(name="Calibri", size=10, color=MID_BLUE, underline="single")
        cell.hyperlink = f"#{name}!A1"
        ws.row_dimensions[i].height = 16

    # A way back to the live page. A workbook is a snapshot the moment it is
    # saved, and the honest thing is to say so and point at the current version
    # rather than let a reader assume a six-month-old file is today's view.
    _live = toc_start + len(sheetnames) + 2
    lc = ws.cell(row=_live, column=2,
                 value=f"↗  Open {ticker} live on QuantWizard (refreshes this analysis)")
    lc.font      = Font(name="Calibri", size=10, bold=True, color=MID_BLUE,
                        underline="single")
    lc.hyperlink = f"{LIVE_BASE}/?page=analysis&ticker={ticker}"
    nc = ws.cell(row=_live + 1, column=2,
                 value="Figures below are as of the generation date above.")
    nc.font = Font(name="Calibri", size=8, italic=True, color="888888")


# ── Dashboard ─────────────────────────────────────────────────────────────────
def _narrative_box(ws, row, text, height=60, italic=True, bold=False, bg=None):
    """Wrapped, full-width (A:D) text box on one tall row. Returns the next row.

    Used for the plain-English blocks (takeaway, what-changed, disclaimers) that
    answer the "dense numbers, little narrative guidance" critique."""
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Calibri", size=10, italic=italic, bold=bold)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border    = _border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[row].height = height
    return row + 1


def _relative_performance_rows(df):
    """(name, ticker_ret, bench_ret, diff) per benchmark present in df.

    Uses the *_Cumulative columns fetch_stock_data already merges in (indexed to
    100 at the period start), so no extra data fetch is needed."""
    rows = []
    if "Cumulative_Index" not in df.columns:
        return rows
    tcum = df["Cumulative_Index"].dropna()
    if len(tcum) < 2:
        return rows
    t_ret = tcum.iloc[-1] / tcum.iloc[0] - 1
    for b, name in [("SPY", "S&P 500 (SPY)"), ("QQQ", "NASDAQ 100 (QQQ)")]:
        col = f"{b}_Cumulative"
        if col in df.columns:
            bcum = df[col].dropna()
            if len(bcum) >= 2:
                b_ret = bcum.iloc[-1] / bcum.iloc[0] - 1
                rows.append((name, t_ret, b_ret, t_ret - b_ret))
    return rows


def _relative_performance_periods(df, bench="SPY"):
    """[(label, stock_ret, bench_ret, diff)] over 1M / 3M / 6M / 1Y / full.

    The whole-period figure above answers "how did this do over five years",
    which hides everything that happened inside it — a stock can beat the index
    over five years while lagging it for the last twelve months. Trading-day
    windows (21/63/126/252) rather than calendar dates, since the frame is
    already indexed by session.
    """
    col = f"{bench}_Cumulative"
    if "Cumulative_Index" not in df.columns or col not in df.columns:
        return []
    t = df["Cumulative_Index"].dropna()
    b = df[col].dropna()
    n = min(len(t), len(b))
    if n < 2:
        return []
    t, b = t.iloc[-n:], b.iloc[-n:]

    out = []
    for label, back in [("1 month", 21), ("3 months", 63), ("6 months", 126),
                        ("1 year", 252), ("Full period", n - 1)]:
        if back < 1 or back >= n:
            continue                      # not enough history for this window
        t0, t1 = float(t.iloc[-1 - back]), float(t.iloc[-1])
        b0, b1 = float(b.iloc[-1 - back]), float(b.iloc[-1])
        if t0 <= 0 or b0 <= 0:
            continue
        tr, br = t1 / t0 - 1, b1 / b0 - 1
        out.append((label, tr, br, tr - br))
    return out


def _technical_posture(df):
    """Descriptive read of the current technical indicators — NOT a recommendation.

    Returns {score:0-100, label, color, signals:[str,...]} or None. The label
    ("Bullish/Mixed/Bearish technicals") describes what the indicators say today;
    it is deliberately framed as a signal, not advice."""
    if df is None or len(df) < 2:
        return None
    latest = df.iloc[-1]

    def _num(col):
        v = latest.get(col)
        try:
            return float(v) if v is not None and pd.notna(v) else None
        except Exception:
            return None

    close, ma50, ma200 = _num("Close"), _num("MA50"), _num("MA200")
    rsi, macd_h, pct_hi = _num("RSI14"), _num("MACD_Hist"), _num("Pct_From_52W_High")
    comps = []   # (points 0..1, descriptive text)

    if close is not None and ma50 is not None:
        comps.append((1.0 if close > ma50 else 0.0,
                      f"Price is {'above' if close > ma50 else 'below'} the 50-day moving average"))
    if close is not None and ma200 is not None:
        comps.append((1.0 if close > ma200 else 0.0,
                      f"Price is {'above' if close > ma200 else 'below'} the 200-day moving average"))
    if ma50 is not None and ma200 is not None:
        up = ma50 > ma200
        comps.append((1.0 if up else 0.0,
                      f"50-day MA is {'above' if up else 'below'} the 200-day MA "
                      f"({'uptrend' if up else 'downtrend'} structure)"))
    if rsi is not None:
        # Five zones, shared with analysis.generate_summary_paragraph — the two
        # once disagreed about the same number ("positive" here, "neutral" there).
        zone = ("overbought" if rsi > 70 else "oversold" if rsi < 30
                else "positive momentum" if rsi > 55
                else "neutral" if rsi >= 45 else "soft momentum")
        comps.append((max(0.0, min(1.0, (rsi - 30) / 40.0)), f"RSI is {rsi:.0f} — {zone}"))
    if macd_h is not None:
        up = macd_h > 0
        comps.append((1.0 if up else 0.0,
                      f"MACD is {'above' if up else 'below'} its signal line "
                      f"({'bullish' if up else 'bearish'} momentum)"))
    if pct_hi is not None:
        comps.append((max(0.0, min(1.0, 1.0 + pct_hi / 0.5)),
                      f"Trading {abs(pct_hi)*100:.0f}% "
                      f"{'below' if pct_hi < 0 else 'above'} the 52-week high"))
    rel = _relative_performance_rows(df)
    if rel:
        name, _t, _b, diff = rel[0]
        short = name.split(" (")[0]
        comps.append((1.0 if diff >= 0 else 0.0,
                      f"{'Outperforming' if diff >= 0 else 'Lagging'} {short} by "
                      f"{abs(diff)*100:.1f} pts over the period"))

    if not comps:
        return None
    score = round(sum(p for p, _ in comps) / len(comps) * 100)
    label, color = (("Bullish technicals", GREEN_OK) if score >= 66 else
                    ("Mixed technicals",   DARK_BLUE) if score >= 40 else
                    ("Bearish technicals",  RED_BAD))
    return {"score": score, "label": label, "color": color,
            "signals": [t for _, t in comps]}


def _recent_changes(df, ticker, lookback=5):
    """Plain-English 'what changed recently' bullets from the last few sessions."""
    if df is None or len(df) < 2:
        return []
    lookback = min(lookback, len(df) - 1)
    latest   = df.iloc[-1]
    bullets  = []

    c_now, c_prev = float(latest["Close"]), float(df["Close"].iloc[-1 - lookback])
    if c_prev:
        bullets.append(f"Price {(c_now / c_prev - 1) * 100:+.1f}% over the last "
                       f"{lookback} sessions (${c_prev:,.2f} -> ${c_now:,.2f}).")

    last_rets = df["Daily_Return"].dropna().tail(lookback)
    if len(last_rets):
        bullets.append(f"Best day {last_rets.max() * 100:+.1f}%, worst day "
                       f"{last_rets.min() * 100:+.1f}% in that window.")

    vv = latest.get("Volume_vs_Avg")
    if vv is not None and pd.notna(vv):
        vv = float(vv)
        if vv >= 1.15 or vv <= 0.85:
            bullets.append(f"Latest volume ran {vv:.1f}x its 20-day average "
                           f"({(vv - 1) * 100:+.0f}%) — {'heavier' if vv >= 1 else 'lighter'} trading.")

    if "RSI14" in df.columns:
        r = df["RSI14"].dropna()
        if len(r) > lookback:
            r_now, r_prev = float(r.iloc[-1]), float(r.iloc[-1 - lookback])
            note = (" — now overbought" if r_now > 70 else
                    " — now oversold" if r_now < 30 else "")
            bullets.append(f"RSI moved {r_prev:.0f} -> {r_now:.0f}{note}.")

    if "Close_vs_MA50" in df.columns:
        cvm = df["Close_vs_MA50"].dropna()
        if len(cvm) > lookback and (cvm.iloc[-1] > 0) != (cvm.iloc[-1 - lookback] > 0):
            above = cvm.iloc[-1] > 0
            bullets.append(f"Price crossed {'above' if above else 'below'} its 50-day "
                           f"moving average in the last {lookback} sessions.")

    if "52W_High" in df.columns:
        hs = df["52W_High"].dropna()
        if len(hs) > lookback and hs.iloc[-1] > hs.iloc[-1 - lookback]:
            bullets.append("Notched a new 52-week high in the last week.")
    if "52W_Low" in df.columns:
        ls = df["52W_Low"].dropna()
        if len(ls) > lookback and ls.iloc[-1] < ls.iloc[-1 - lookback]:
            bullets.append("Made a new 52-week low in the last week.")

    if "SPY_Cumulative" in df.columns:
        s = df["SPY_Cumulative"].dropna()
        t = df["Cumulative_Index"].dropna()
        if len(s) > lookback and len(t) > lookback:
            diff = (t.iloc[-1] / t.iloc[-1 - lookback] - 1) - (s.iloc[-1] / s.iloc[-1 - lookback] - 1)
            bullets.append(f"{'Outpaced' if diff >= 0 else 'Trailed'} the S&P 500 by "
                           f"{abs(diff) * 100:.1f} pts over the last {lookback} sessions.")

    return bullets[:5]


def _coerce_metric(v):
    """('25.6%', '43.14', 1000) → typed cell value + number format.

    The Monte Carlo summary arrives as display strings; writing them verbatim
    left a column of numbers-as-text (left-aligned, green error triangles,
    unusable in formulas). Anything that doesn't parse cleanly stays text.
    """
    if isinstance(v, bool) or v is None:
        return v, None
    if isinstance(v, (int, float)):
        return v, ("0" if float(v).is_integer() and abs(v) >= 100 else "0.00")
    s = str(v).strip()
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", "")) / 100, "0.0%"
        except ValueError:
            return v, None
    try:
        f = float(s.replace(",", "").replace("$", ""))
    except ValueError:
        return v, None
    if f.is_integer() and abs(f) >= 100:      # counts: horizon days, simulations
        return int(f), "0"
    return f, '"$"#,##0.00'                   # everything else here is a price


def _mc_plain_language(mc_summary):
    """Turn the Monte Carlo summary into plain-English downside/upside lines."""
    if not mc_summary:
        return []
    last    = mc_summary.get("Last Price")
    horizon = mc_summary.get("Forecast Horizon (days)")
    lines   = []
    if last:
        lines.append(f"From today's ${last:,.2f}"
                     + (f" over ~{horizon} trading days:" if horizon else ":"))

    def _line(label, key):
        v = mc_summary.get(key)
        if v is None:
            return
        chg = f" ({(v / last - 1) * 100:+.1f}% vs today)" if last else ""
        lines.append(f"{label}: ${v:,.2f}{chg}")

    _line("Median outcome (P50)",         "Median (P50)")
    _line("Bear case (P5, ~5% chance)",   "Bear Case (P5)")
    _line("Low case (P25, ~25% chance)",  "Low Case (P25)")
    _line("Bull case (P75, ~25% chance)", "Bull Case (P75)")
    _line("Best case (P95, ~5% chance)",  "Best Case (P95)")
    prob = mc_summary.get("Prob. of Gain")
    if prob is not None:
        lines.append(f"Probability of finishing above today's price: {prob}.")
    return lines


def _humanize_company_value(key, value):
    """Format company_details values for display instead of dumping raw floats.

    Fixes the "Market Cap 3288094076962.04" issue flagged in review — large money
    magnitudes render as $T/$B/$M and plain counts get thousands separators.
    Non-numeric values (already-formatted strings) pass through untouched."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return str(value)
    k = key.lower()
    if any(w in k for w in ("cap", "value", "revenue", "assets", "debt", "cash")):
        a = abs(value)
        if a >= 1e12: return f"${value / 1e12:,.2f}T"
        if a >= 1e9:  return f"${value / 1e9:,.2f}B"
        if a >= 1e6:  return f"${value / 1e6:,.2f}M"
        return f"${value:,.0f}"
    if float(value).is_integer():
        return f"{int(value):,}"
    return f"{value:,.2f}"


def _build_dashboard(wb, ticker, df, company_details, mc_summary,
                     resistance_levels, support_levels, summary_text,
                     analyst_data=None, dcf=None, fundamentals=None):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10

    latest     = df.iloc[-1]
    first      = df.iloc[0]
    period_ret = (latest["Close"] / first["Close"] - 1) * 100

    ret      = df["Daily_Return"].dropna()
    ann_ret  = ret.mean() * 252
    ann_std  = ret.std() * np.sqrt(252)
    # Excess-return Sharpe/Sortino (subtract the risk-free rate) so the exported
    # report matches the on-screen metric cards and the portfolio engine. The
    # Sortino denominator is downside deviation about zero, not the std of the
    # losing days — see analysis.downside_deviation.
    from analysis import downside_deviation as _dd_fn
    downside = _dd_fn(ret)
    rfr      = get_risk_free_rate()
    sharpe   = (ann_ret - rfr) / ann_std  if ann_std  else np.nan
    sortino  = (ann_ret - rfr) / downside if downside else np.nan

    try:
        rsi_val = float(latest.get("RSI14", np.nan))
    except Exception:
        rsi_val = np.nan

    ws.merge_cells("A1:D1")
    ws["A1"] = f"{ticker} — Equity Research Dashboard"
    ws["A1"].font      = Font(size=18, bold=True, color=DARK_BLUE, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:D2")
    ws["A2"] = (f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
                f"  |  Multi-source data: {DATA_SOURCE_LINE}")
    ws["A2"].font      = Font(italic=True, color="888888", name="Calibri", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    def sec_hdr(row, label, col_end="D"):
        ws.merge_cells(f"A{row}:{col_end}{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, color=WHITE, name="Calibri", size=11)
        c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    def kv(row, label, value, fmt=None, rag=None):
        cl = ws.cell(row=row, column=1, value=label)
        cv = ws.cell(row=row, column=2, value=value)
        cl.font      = Font(name="Calibri", size=10)
        cv.font      = Font(name="Calibri", size=10, bold=True)
        cv.alignment = Alignment(horizontal="right")
        cl.border    = cv.border = _border()
        if fmt and isinstance(value, (int, float)):
            cv.number_format = fmt
        if rag and isinstance(value, (int, float)):
            direction, thresh = rag
            # Neutral band around a zero threshold: -0.7% vs the 50-day MA was
            # getting the same solid-red alarm as -72%. Within ±2% (or ±2 pts
            # for point-spread cells) the cell stays unfilled.
            band = 2.0 if (fmt and "pts" in fmt) else 0.02
            if direction == "gt" and thresh == 0 and abs(value) <= band:
                pass
            else:
                colour = (GREEN_OK if value > thresh else RED_BAD) if direction == "gt" \
                         else (RED_BAD if value < thresh else GREEN_OK)
                cv.fill = PatternFill("solid", fgColor=colour)
                cv.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        bg = GREY_ROW if row % 2 == 0 else WHITE
        for col in [1, 2]:
            c = ws.cell(row=row, column=col)
            if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE):
                c.fill = PatternFill("solid", fgColor=bg)

    # Plain currency, not the accounting format: in a 24-wide column the
    # accounting '$' floats an inch left of the digits and reads unfinished.
    USD = '"$"#,##0.00'

    sec_hdr(4, "Price & Performance")
    kv(5,  "Current Price ($)",        latest["Close"],              fmt=USD)
    kv(6,  "Period Return",            period_ret / 100,             fmt="0.00%", rag=("gt", 0))
    kv(7,  "52-Week High ($)",         latest.get("52W_High"),       fmt=USD)
    kv(8,  "52-Week Low ($)",          latest.get("52W_Low"),        fmt=USD)
    kv(9,  "% from 52W High",          latest.get("Pct_From_52W_High"), fmt="0.00%", rag=("gt", -0.10))
    kv(10, "20-Day MA ($)",            latest.get("MA20"),           fmt=USD)
    kv(11, "50-Day MA ($)",            latest.get("MA50"),           fmt=USD)
    kv(12, "200-Day MA ($)",           latest.get("MA200"),          fmt=USD)
    kv(13, "Price vs 50-Day MA",       latest.get("Close_vs_MA50"),  fmt="0.00%", rag=("gt", 0))

    if "BB_Upper" in df.columns:
        sec_hdr(15, "Bollinger Bands (20-day, 2σ)")
        kv(16, "BB Upper ($)",  latest.get("BB_Upper"),  fmt=USD)
        kv(17, "BB Middle ($)", latest.get("BB_Middle"), fmt=USD)
        kv(18, "BB Lower ($)",  latest.get("BB_Lower"),  fmt=USD)
        kv(19, "BB Width",      latest.get("BB_Width"),  fmt="0.0000")
        kv(20, "BB %B",         latest.get("BB_Pct"),    fmt="0.00%")
        risk_start = 22
    else:
        risk_start = 15

    sec_hdr(risk_start, "Risk & Return Metrics")
    kv(risk_start+1, "20-Day Ann. Volatility",  latest.get("Volatility_20d"), fmt="0.00%")
    kv(risk_start+2, "60-Day Max Drawdown",      df["Drawdown_60d"].min(),     fmt="0.00%", rag=("gt", -0.20))
    kv(risk_start+3, "RSI (14)",
       round(rsi_val, 1) if pd.notna(rsi_val) else "N/A",
       rag=("lt", 70) if pd.notna(rsi_val) and rsi_val > 70 else
           ("gt", 30) if pd.notna(rsi_val) and rsi_val < 30 else None)
    kv(risk_start+4, "Sharpe Ratio",  round(sharpe, 2)  if pd.notna(sharpe)  else "N/A", rag=("gt", 1))
    kv(risk_start+5, "Sortino Ratio", round(sortino, 2) if pd.notna(sortino) else "N/A", rag=("gt", 1))

    row_cursor = risk_start + 7

    # ── Investor Takeaway — a plain-English lead so the page opens with a view,
    #    not a wall of numbers (the "dense, little narrative guidance" critique).
    posture   = _technical_posture(df)
    consensus = consensus_from_recommendation((analyst_data or {}).get("recommendation"))
    changes   = _recent_changes(df, ticker)

    sec_hdr(row_cursor, "Investor Takeaway")
    row_cursor += 1
    takeaway = [f"{ticker} is {period_ret:+.1f}% over the period, last ${latest['Close']:,.2f}."]
    if posture:
        takeaway.append(f"Technical posture: {posture['label']} ({posture['score']}/100).")
    if consensus:
        takeaway.append(f"Wall-Street consensus: {consensus['verdict']} "
                        f"({consensus['total']} analysts).")
    if dcf and dcf.get("ok") and dcf.get("upside") is not None:
        takeaway.append(f"DCF fair value ${dcf['fair_value']:,.0f} "
                        f"({dcf['upside']*100:+.0f}% vs price).")
    row_cursor = _narrative_box(ws, row_cursor, "  ".join(takeaway),
                                height=58, italic=False, bold=True, bg=TILE_BG)

    # ── Stock Scorecard — the "investment snapshot" that turns a dozen metrics
    #    into a graded profile. Descriptive (quality/attractiveness), not buy/sell.
    # The scorecard's risk bands (-0.10 -> 85 ... -0.50 -> 30) are calibrated for a
    # true peak-to-trough max drawdown, and compute_scorecard labels the value
    # "Max DD". Feeding it Drawdown_60d — the worst 60-DAY rolling drawdown — was
    # handing it a systematically shallower number, so the Risk factor scored
    # every stock more kindly than the bands intend.
    _sc_cum = (1 + ret).cumprod()
    _sc_mdd = float((_sc_cum / _sc_cum.cummax() - 1).min()) if len(_sc_cum) else None
    _sc_risk = {
        "sharpe": float(sharpe)  if pd.notna(sharpe)  else None,
        "vol":    float(ann_std) if pd.notna(ann_std) else None,
        "max_dd": _sc_mdd if (_sc_mdd is not None and pd.notna(_sc_mdd)) else None,
    }
    scorecard = compute_scorecard(
        fundamentals=fundamentals, dcf=dcf,
        momentum_score=(posture["score"] if posture else None),
        risk=_sc_risk, consensus=consensus)
    if scorecard and scorecard.get("ok"):
        _grade_bg = {"Strong": "548235", "Above-avg": GREEN_OK, "Average": "BF8F00",
                     "Below-avg": "C55A11", "Weak": RED_BAD}
        sec_hdr(row_cursor, "Stock Scorecard")
        row_cursor += 1
        comp  = scorecard["composite"]
        ccolor = GREEN_OK if comp >= 65 else "BF8F00" if comp >= 45 else RED_BAD
        cl = ws.cell(row=row_cursor, column=1, value="Composite Score (0–100)")
        cv = ws.cell(row=row_cursor, column=2, value=comp)
        cl.font = Font(name="Calibri", size=10, bold=True)
        cv.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        cv.fill = PatternFill("solid", fgColor=ccolor)
        cv.alignment = Alignment(horizontal="right")
        cl.border = cv.border = _border()
        row_cursor += 1
        ll = ws.cell(row=row_cursor, column=1, value="Overall Profile")
        ws.merge_cells(f"B{row_cursor}:C{row_cursor}")
        lv = ws.cell(row=row_cursor, column=2, value=scorecard["label"])
        ll.font = Font(name="Calibri", size=10)
        lv.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        lv.fill = PatternFill("solid", fgColor=ccolor)
        lv.alignment = Alignment(horizontal="right")
        ll.border = lv.border = _border()
        row_cursor += 1
        for ci, h in enumerate(["Factor", "Score", "Grade"], 1):
            _hdr_cell(ws.cell(row=row_cursor, column=ci, value=h), bg=MID_BLUE)
        row_cursor += 1
        for fac in scorecard["factors"]:
            a = ws.cell(row=row_cursor, column=1, value=fac["name"])
            b = ws.cell(row=row_cursor, column=2, value=fac["score"])
            c = ws.cell(row=row_cursor, column=3, value=fac["grade"])
            a.font = Font(name="Calibri", size=10)
            b.font = Font(name="Calibri", size=10, bold=True)
            b.alignment = Alignment(horizontal="right")
            c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=_grade_bg.get(fac["grade"], MID_BLUE))
            c.alignment = Alignment(horizontal="center")
            a.border = b.border = c.border = _border()
            row_cursor += 1
        _details = "   ·   ".join(f"{fac['name']}: {fac['detail']}"
                                   for fac in scorecard["factors"] if fac["detail"])
        row_cursor = _narrative_box(
            ws, row_cursor,
            "Weighted blend of valuation, growth, profitability, financial health, "
            "momentum, risk and sentiment — a descriptive profile of the stock's "
            f"characteristics, not a recommendation.\n{_details}\n" + DISCLAIMER_SHORT,
            height=96, italic=True, bg="FFF8E1")
        row_cursor += 1

    if changes:
        c = ws.cell(row=row_cursor, column=1, value="What changed recently")
        c.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
        row_cursor += 1
        bullet_text = "\n".join(f"•  {b}" for b in changes)
        est_lines   = sum(max(1, math.ceil(len(b) / 82)) for b in changes)
        row_cursor  = _narrative_box(ws, row_cursor, bullet_text,
                                     height=16 * est_lines + 8, italic=False)
    row_cursor += 1

    # ── Relative Performance vs benchmarks (uses *_Cumulative already in df) ────
    rel_rows = _relative_performance_rows(df)
    if rel_rows:
        sec_hdr(row_cursor, "Relative Performance")
        row_cursor += 1
        t_ret = rel_rows[0][1]
        for name, _t, _b, diff in rel_rows:
            # The gap between two cumulative returns is percentage POINTS, not a
            # return — "-151.8%" formatted as a percent reads as an impossible loss.
            kv(row_cursor, f"vs {name} (difference)", diff * 100,
               fmt='+0.0" pts";-0.0" pts"', rag=("gt", 0))
            row_cursor += 1
        abs_bits = [f"{ticker} {t_ret*100:+.1f}%"] + \
                   [f"{n.split(' (')[0]} {b*100:+.1f}%" for n, _t, b, _d in rel_rows]
        row_cursor = _narrative_box(ws, row_cursor, "Period return:    " + "      |      ".join(abs_bits),
                                    height=26, italic=False)
        row_cursor += 1

        # Broken out by window: the single whole-period number hides a stock that
        # beat the index over five years while lagging it over the last twelve
        # months, which is usually the more decision-relevant fact.
        periods = _relative_performance_periods(df)
        if periods:
            # "(pts)" in the header, not only in the cell's number format. The
            # first two columns hold fractions and the third holds percentage
            # points, so the values are on different scales even though all three
            # render correctly on screen. Anyone reading the cells rather than
            # the sheet — a script, or a reviewer — sees 0.0338 beside 5.08 and
            # reasonably concludes the units are mixed and unmarked.
            hdr = ["Window", ticker, "S&P 500", "Relative (pts)"]
            for ci, htxt in enumerate(hdr, 1):
                c = ws.cell(row=row_cursor, column=ci, value=htxt)
                c.font   = Font(name="Calibri", size=9, bold=True, color=WHITE)
                c.fill   = PatternFill("solid", fgColor=DARK_BLUE)
                c.border = _border()
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            row_cursor += 1
            for label, tr, br, diff in periods:
                vals = [label, tr, br, diff * 100]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=row_cursor, column=ci, value=v)
                    c.font   = Font(name="Calibri", size=10,
                                    bold=(ci == 4),
                                    color=(GREEN_OK if ci == 4 and diff >= 0
                                           else RED_BAD if ci == 4 else "000000"))
                    c.border = _border()
                    if ci > 1:
                        c.number_format = ('+0.0" pts";-0.0" pts"' if ci == 4
                                           else "+0.0%;-0.0%")
                        c.alignment = Alignment(horizontal="right")
                row_cursor += 1
            row_cursor += 1

    # ── Technical Posture — descriptive read of the indicators, NOT advice ──────
    if posture:
        sec_hdr(row_cursor, "Technical Posture")
        row_cursor += 1
        kv(row_cursor, "Technical Score (0–100)", posture["score"])
        row_cursor += 1
        lc = ws.cell(row=row_cursor, column=1, value="Reading")
        vc = ws.cell(row=row_cursor, column=2, value=posture["label"])
        lc.font = Font(name="Calibri", size=10)
        vc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        vc.fill = PatternFill("solid", fgColor=posture["color"])
        vc.alignment = Alignment(horizontal="right")
        lc.border = vc.border = _border()
        row_cursor += 1
        for sig in posture["signals"]:
            sc = ws.cell(row=row_cursor, column=1, value=f"•  {sig}")
            ws.merge_cells(f"A{row_cursor}:D{row_cursor}")
            sc.font = Font(name="Calibri", size=9)
            sc.alignment = Alignment(wrap_text=True, vertical="center")
            row_cursor += 1
        row_cursor = _narrative_box(
            ws, row_cursor,
            "Describes what the technical indicators say today — a signal, not a "
            "recommendation. " + DISCLAIMER_SHORT,
            height=40, italic=True, bg="FFF8E1")
        row_cursor += 1

    # ── Analyst Consensus — Wall Street's view, clearly attributed (not ours) ───
    if consensus:
        sec_hdr(row_cursor, "Analyst Consensus")
        row_cursor += 1
        lc = ws.cell(row=row_cursor, column=1, value="Wall-Street Rating")
        vc = ws.cell(row=row_cursor, column=2, value=consensus["verdict"])
        lc.font = Font(name="Calibri", size=10)
        vc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        vc.fill = PatternFill("solid", fgColor=consensus["color"].lstrip("#"))
        vc.alignment = Alignment(horizontal="right")
        lc.border = vc.border = _border()
        row_cursor += 1
        kv(row_cursor, "Buy (Strong Buy + Buy)", consensus["strong_buy"] + consensus["buy"])
        row_cursor += 1
        kv(row_cursor, "Hold", consensus["hold"])
        row_cursor += 1
        kv(row_cursor, "Sell (Sell + Strong Sell)", consensus["sell"] + consensus["strong_sell"])
        row_cursor += 1
        kv(row_cursor, "Analysts covering", consensus["total"])
        row_cursor += 1
        row_cursor = _narrative_box(
            ws, row_cursor,
            f"Source: {consensus['total']} Wall-Street analyst ratings aggregated by "
            f"Finnhub (period {consensus['period']}). This is analysts' consensus, "
            f"not QuantWizard's opinion. " + DISCLAIMER_SHORT,
            height=40, italic=True, bg="FFF8E1")
        row_cursor += 1

    # ── Valuation — leads with the reverse DCF (what the price already assumes),
    #    then the forward conclusion. Full editable model on the Valuation sheet.
    if dcf and dcf.get("ok"):
        sec_hdr(row_cursor, "Valuation — Reverse DCF & Fair Value")
        row_cursor += 1
        up  = dcf.get("upside")
        imp = dcf.get("market_implied_growth")
        verdict, vcolor = (("Undervalued vs DCF", GREEN_OK) if up is not None and up > 0.15 else
                           ("Overvalued vs DCF",  RED_BAD)  if up is not None and up < -0.15 else
                           ("Fairly valued vs DCF", DARK_BLUE))
        if imp is not None:
            kv(row_cursor, "Market-Implied FCF Growth", imp, fmt="0.0%")
        else:
            kv(row_cursor, "Market-Implied FCF Growth", "Outside solvable range")
        row_cursor += 1
        kv(row_cursor, "DCF Fair Value / Share", dcf["fair_value"], fmt='"$"#,##0.00')
        row_cursor += 1
        if up is not None:
            kv(row_cursor, "Upside / Downside", up, fmt="+0.0%;-0.0%", rag=("gt", 0))
            row_cursor += 1
        lc = ws.cell(row=row_cursor, column=1, value="DCF Verdict")
        vc = ws.cell(row=row_cursor, column=2, value=verdict)
        lc.font = Font(name="Calibri", size=10)
        vc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        vc.fill = PatternFill("solid", fgColor=vcolor)
        vc.alignment = Alignment(horizontal="right")
        lc.border = vc.border = _border()
        row_cursor += 1
        scn  = dcf.get("scenarios", {})
        bear, bull = scn.get("bear", {}), scn.get("bull", {})
        if imp is not None:
            note = (f"Today's ${dcf['price']:,.2f} price implies free cash flow "
                    f"compounding at ~{imp*100:.1f}% a year for {dcf['years']} years. "
                    f"That is the number to argue with. ")
        else:
            note = (f"Today's ${dcf['price']:,.2f} price sits outside the growth range "
                    f"the reverse DCF can solve, so no implied rate is quoted. ")
        note += (f"2-stage FCF DCF at a {dcf['wacc']*100:.1f}% discount rate, "
                 f"{dcf['terminal_growth']*100:.1f}% terminal growth, "
                 f"base case {dcf['base_growth']*100:.1f}% stage-1 growth. ")
        if bear.get("fair_value") and bull.get("fair_value"):
            note += f"Bear ${bear['fair_value']:,.0f} / Bull ${bull['fair_value']:,.0f}. "
        note += ("The Valuation sheet holds the full model as live formulas — change an "
                 "assumption there and it recalculates. ") + DISCLAIMER_SHORT
        row_cursor = _narrative_box(ws, row_cursor, note, height=70, italic=True, bg="FFF8E1")
        row_cursor += 1

        # Conditions the fair value rests on, next to the fair value rather than
        # in the Methodology sheet. dcf_valuation raises these when the filer is
        # a bank or broker (where unlevered FCF is the wrong instrument) or when
        # the base was built by discarding negative years. On a red ground,
        # because they qualify the number immediately above them.
        for _cav in (dcf.get("caveats") or []):
            row_cursor = _narrative_box(ws, row_cursor, _cav, height=56,
                                        italic=False, bg="FDECEA")
        if dcf.get("caveats"):
            row_cursor += 1
    else:
        sec_hdr(row_cursor, "Valuation — Reverse DCF & Fair Value")
        row_cursor += 1
        _reason = (dcf or {}).get("reason", "") if isinstance(dcf, dict) else ""
        row_cursor = _narrative_box(
            ws, row_cursor,
            f"No discounted-cash-flow model for {ticker}.  "
            + _DCF_NO_MODEL_REASONS.get(
                _reason,
                "The inputs a DCF needs — audited financial statements, a market price "
                "and a positive free-cash-flow history — were not all available.")
            + "  " + DISCLAIMER_SHORT,
            height=56, italic=True, bg="FFF8E1")
        row_cursor += 1

    if resistance_levels or support_levels:
        sec_hdr(row_cursor, "Support & Resistance Levels")
        row_cursor += 1

        # Levels get their own left-aligned merged row: five joined prices
        # right-aligned in column B clipped the leading levels clean off.
        def _levels_row(row, label, levels):
            lc = ws.cell(row=row, column=1, value=label)
            lc.font, lc.border = Font(name="Calibri", size=10), _border()
            ws.merge_cells(f"B{row}:D{row}")
            vc = ws.cell(row=row, column=2,
                         value="   |   ".join(f"${v:,.2f}" for v in levels) if levels else "—")
            vc.font      = Font(name="Calibri", size=10, bold=True)
            vc.alignment = Alignment(horizontal="left")
            vc.border    = _border()

        _levels_row(row_cursor,     "Resistance (above price)", resistance_levels or [])
        _levels_row(row_cursor + 1, "Support (below price)",    support_levels or [])
        row_cursor += 2
        ws.merge_cells(f"A{row_cursor}:D{row_cursor}")
        cap = ws.cell(row=row_cursor, column=1,
                      value="Swing highs/lows from the last 12 months, nearest to "
                            "the current price first.")
        cap.font = Font(name="Calibri", size=8, italic=True, color="888888")
        row_cursor += 2

    if mc_summary:
        sec_hdr(row_cursor, "Monte Carlo Forecast")
        row_cursor += 1
        for k, v in mc_summary.items():
            val, vfmt = _coerce_metric(v)
            kv(row_cursor, k, val, fmt=vfmt)
            row_cursor += 1
        mc_lines = _mc_plain_language(mc_summary)
        if mc_lines:
            body = ("Monte Carlo in plain English\n" + "\n".join(mc_lines)
                    + "\nSimulated from the historical return distribution — outcomes "
                      "are probabilities, not guarantees.")
            row_cursor = _narrative_box(ws, row_cursor, body,
                                        height=16 * (len(mc_lines) + 3) + 6,
                                        italic=False, bg=TILE_BG)
        row_cursor += 1

    if company_details:
        sec_hdr(row_cursor, "Company Information")
        row_cursor += 1
        for k, v in company_details.items():
            if k == "Description":
                continue
            # Polygon has one SIC description, not a sector/industry pair —
            # printing it twice under two labels reads as a data glitch.
            if k == "Industry" and v == company_details.get("Sector"):
                continue
            kv(row_cursor, k, _humanize_company_value(k, v))
            row_cursor += 1
        row_cursor += 1

    sec_hdr(row_cursor, "Automated Analysis Summary", col_end="D")
    row_cursor += 1
    ws.merge_cells(f"A{row_cursor}:D{row_cursor + 4}")
    sc = ws.cell(row=row_cursor, column=1, value=summary_text)
    sc.font      = Font(name="Calibri", size=10, italic=True)
    sc.alignment = Alignment(wrap_text=True, vertical="top")
    sc.border    = _border()
    ws.row_dimensions[row_cursor].height = 90
    row_cursor += 6

    if MPL_AVAILABLE:
        ws.column_dimensions["E"].width = 18
        spark_data = [
            ("Price",      df["Close"].tolist(),          "#2E75B6"),
            ("Volume",     df["Volume"].tolist(),         "#70AD47"),
            ("Daily Ret",  df["Daily_Return"].tolist(),   "#FF6B35"),
            ("Volatility", df["Volatility_20d"].tolist(), "#7030A0"),
            ("Drawdown",   df["Drawdown_60d"].tolist(),   "#C00000"),
        ]
        ws.cell(row=4, column=5, value="SPARKLINES").font = Font(bold=True, color=WHITE, name="Calibri")
        ws.cell(row=4, column=5).fill = PatternFill("solid", fgColor=MID_BLUE)
        for i, (label, vals, col) in enumerate(spark_data):
            row = 5 + i * 3
            ws.cell(row=row, column=5, value=label).font = Font(name="Calibri", size=9, bold=True)
            buf = make_sparkline(vals, color=col)
            if buf:
                img = XLImage(buf)
                img.width, img.height = 130, 35
                ws.add_image(img, f"E{row+1}")
            # No row-height bump: images float over the grid, and stretching
            # rows 6/9/12/15/18 made every third KV row read double-height.

    return ws


# ── Annual summary sheet ──────────────────────────────────────────────────────
def _build_annual_summary(wb, df):
    """Year-by-year performance table — always included, especially useful for long ranges."""
    ws_a = wb.create_sheet("Annual_Summary")
    ws_a.sheet_view.showGridLines = False

    ws_a.merge_cells("A1:H1")
    ws_a["A1"] = "Annual Performance Summary"
    ws_a["A1"].font      = Font(size=14, bold=True, color=DARK_BLUE, name="Calibri")
    ws_a["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_a.row_dimensions[1].height = 26

    headers = ["Year","Annual Return","Max Drawdown (60d)","Ann. Volatility","Sharpe"]
    for ci, h in enumerate(headers, 1):
        _hdr_cell(ws_a.cell(row=2, column=ci, value=h), bg=MID_BLUE)

    tmp = df[["Date","Daily_Return","Drawdown_60d"]].copy()
    tmp["Year"] = pd.to_datetime(tmp["Date"]).dt.year

    # This was the one Sharpe in the app that did NOT net off the risk-free rate,
    # so the per-year column read high by roughly Rf/vol against the Dashboard's
    # figure and against this workbook's own Methodology sheet.
    _rfr = get_risk_free_rate()

    # A partial current year printed as a bare "2026" reads as a full annual
    # return; a YTD marker keeps the -31% honest.
    _last_dt = pd.to_datetime(tmp["Date"]).max()

    for ri, (year, grp) in enumerate(tmp.groupby("Year"), 3):
        ret         = grp["Daily_Return"].dropna()
        yr_return   = (1 + ret).prod() - 1
        yr_drawdown = grp["Drawdown_60d"].min()
        yr_vol      = ret.std() * np.sqrt(252)
        yr_sharpe   = ((ret.mean() * 252) - _rfr) / yr_vol if yr_vol else np.nan

        _ytd = (year == _last_dt.year
                and not (_last_dt.month == 12 and _last_dt.day >= 28))
        row_vals = [f"{year} (YTD)" if _ytd else year, yr_return, yr_drawdown,
                    yr_vol, round(yr_sharpe, 2) if pd.notna(yr_sharpe) else "N/A"]
        bg = GREY_ROW if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row_vals, 1):
            c = ws_a.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Calibri", size=10)
            c.border = _border()
            c.fill   = PatternFill("solid", fgColor=bg)
            if ci == 1: c.number_format = "0"
            elif ci in (2, 3, 4):
                c.number_format = "0.00%"
                # RAG only for return (col 2) and drawdown (col 3); volatility (col 4)
                # is not "good/bad" on its own, so it stays neutral.
                if isinstance(val, float) and ci in (2, 3):
                    good = val > 0 if ci == 2 else val > -0.15
                    c.fill = PatternFill("solid", fgColor=GREEN_OK if good else BAD_FILL)
                    c.font = Font(name="Calibri", size=10, bold=True,
                                  color=WHITE if good else BAD_TEXT)

    auto_col_width(ws_a)
    ws_a.freeze_panes = "A3"
    ws_a.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    return ws_a


# ── Price & Indicators sheet ──────────────────────────────────────────────────
def _build_price_sheet(wb, df, bar_size="day"):
    # This sheet was 30 columns and did not include the price. It carried MA20,
    # Close_vs_MA20/50/200, four Bollinger columns and three MACD columns, but no
    # Open/High/Low/Close/Volume — so you could read "2% above the 20-day MA"
    # without being able to see what the stock actually closed at. BB_Middle was
    # also a duplicate: the middle Bollinger band IS the 20-day moving average.
    #
    # Kept: the raw data, plus one column per question rather than one per
    # formula. MACD_Hist is the signal (it already encodes MACD vs its own
    # signal line); the 52-week percentages are what gets read, not the raw
    # high/low; MA20 is recoverable from the price and rarely traded off.
    price_cols = [c for c in
                  ["Date", "Open", "High", "Low", "Close", "Volume",
                   "Daily_Return", "Cumulative_Index",
                   "MA50", "MA200",
                   "Volatility_20d", "Drawdown_60d", "Pct_From_52W_High"]
                  if c in df.columns]
    if "RSI14" in df.columns:
        price_cols += [c for c in ["RSI14", "MACD_Hist"] if c in df.columns]
    # Benchmark comparison stays — it's the one thing here a price chart can't show.
    price_cols += [c for c in df.columns if c.endswith("_Cumulative")]

    full_df   = df[[c for c in price_cols if c in df.columns]].copy()
    # Cap raw data sheet at 1,300 rows (~5yr daily). All calculations use the full dataset.
    ROW_CAP   = 1300
    truncated = len(full_df) > ROW_CAP
    export_df = full_df.tail(ROW_CAP).copy() if truncated else full_df

    ws_p = wb.create_sheet("Price_Indicators")

    # Info banner when data is capped
    if truncated:
        ws_p.insert_rows(1)
        note = (f"Note: Showing most recent {ROW_CAP} bars ({bar_size} data) to keep "
                f"the workbook a readable size — this is a deliberate display cap, "
                f"not a spreadsheet row limit. "
                f"Full {len(full_df)}-bar history used for all calculations & charts. "
                f"See Annual_Summary sheet for full year-by-year breakdown.")
        ws_p.merge_cells(f"A1:{get_column_letter(len(export_df.columns))}1")
        nc = ws_p["A1"]
        nc.value     = note
        nc.font      = Font(name="Calibri", size=9, italic=True, color="1F4E79")
        nc.fill      = PatternFill("solid", fgColor="D6E4F0")
        nc.alignment = Alignment(wrap_text=True, vertical="center")
        ws_p.row_dimensions[1].height = 30

    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws_p.append(r)

    # Header row is row 2 if banner exists, else row 1
    hdr_row = 2 if truncated else 1
    data_start = hdr_row + 1
    for cell in ws_p[hdr_row]:
        _hdr_cell(cell)
    auto_col_width(ws_p)
    ws_p.freeze_panes = f"A{data_start}"
    ws_p.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ws_p.max_column)}{hdr_row}"

    col_map    = {c[0].column_letter: c[0].value
                  for c in ws_p.iter_cols(1, ws_p.max_column, hdr_row, hdr_row)}
    price_hdrs = {"Open","High","Low","Close",
                  "MA20","MA50","MA200","BB_Upper","BB_Middle","BB_Lower",
                  "52W_High","52W_Low"}
    pct_hdrs   = {"Daily_Return","Close_vs_MA20","Close_vs_MA50","Close_vs_MA200",
                  "Volatility_20d","Drawdown_20d","Drawdown_60d","BB_Pct","Volume_vs_Avg",
                  "Pct_From_52W_High","Pct_From_52W_Low"}

    for row in ws_p.iter_rows(min_row=data_start):
        for cell in row:
            h = col_map.get(cell.column_letter)
            if   h == "Date":              cell.number_format = "yyyy-mm-dd"
            elif h in ("Volume", "Vol_MA20"):  cell.number_format = "#,##0"
            elif h == "Cumulative_Index":  cell.number_format = "#,##0.00"  # base-100 index, not a %
            elif h in price_hdrs:          cell.number_format = '_($* #,##0.00_)'
            elif h in pct_hdrs:            cell.number_format = "0.00%"

    dr_col = next((l for l, h in col_map.items() if h == "Daily_Return"), None)
    if dr_col:
        ws_p.conditional_formatting.add(
            f"{dr_col}{data_start}:{dr_col}{ws_p.max_row}",
            ColorScaleRule(start_type="num", start_value=-0.05, start_color="FFAAAA",
                           mid_type="num",   mid_value=0,        mid_color="FFFFFF",
                           end_type="num",   end_value=0.05,     end_color="AAFFAA"))
    rsi_col = next((l for l, h in col_map.items() if h == "RSI14"), None)
    if rsi_col:
        rng = f"{rsi_col}{data_start}:{rsi_col}{ws_p.max_row}"
        ws_p.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["70"],
            fill=PatternFill("solid", fgColor="FF9999")))
        ws_p.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["30"],
            fill=PatternFill("solid", fgColor="99FF99")))
    return ws_p, export_df


# ── News sheet ────────────────────────────────────────────────────────────────
def _build_news_sheet(wb, news_list):
    if not news_list:
        return
    ws_n = wb.create_sheet("News_Headlines")
    # Theme comes from news_research's classifier — it's what makes this read
    # as analysed rather than scraped. Absent on the legacy path, hence .get().
    cols = ["Date", "Headline", "Publisher", "Theme", "Sentiment", "Relevance", "URL"]
    ws_n.append(cols)
    style_header_row(ws_n)
    sent_fill = {"Positive": ("C6EFCE", "006100"), "Negative": ("FFC7CE", "9C0006"),
                 "Neutral":  ("FFF2CC", "7F6000")}

    # Newest first. The feed arrives as whatever order the sources came back in,
    # which produced a sheet with 07-30 above 08-25 and the sequence restarting
    # partway down — two result sets appended without a merge. Unparseable dates
    # sort last rather than crashing the sort.
    def _news_dt(item):
        try:
            v = pd.to_datetime(item.get("Date"))
            return None if pd.isna(v) else v.to_pydatetime()
        except Exception:
            return None
    news_list = sorted(news_list, key=lambda it: (_news_dt(it) or datetime.min),
                       reverse=True)

    for ni, item in enumerate(news_list, 2):
        for ci, key in enumerate(cols, 1):
            c = ws_n.cell(row=ni, column=ci, value=item.get(key, ""))
            # A real datetime with a date format, not text. As a string the
            # column will not sort, filter or feed a date axis — and the first
            # thing anyone does with a news table is sort it by date.
            if key == "Date":
                _d = _news_dt(item)
                if _d is not None:
                    c.value = _d
                    c.number_format = "yyyy-mm-dd"
            c.font   = Font(name="Calibri", size=10)
            c.border = _border()
            if ni % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GREY_ROW)
            # Colour the Sentiment cell so the feed reads as analysed, not raw.
            if key == "Sentiment" and item.get("Sentiment") in sent_fill:
                bg, fg = sent_fill[item["Sentiment"]]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Calibri", size=10, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center")
            if key == "Relevance":
                c.alignment = Alignment(horizontal="center")
            # A real hyperlink, not a column of raw 120-char URLs.
            if key == "URL":
                url = item.get("URL") or ""
                if url:
                    c.value     = "Open article"
                    c.hyperlink = url
                    c.font      = Font(name="Calibri", size=10, color=MID_BLUE,
                                       underline="single")
                else:
                    c.value = ""
    for col, w in zip("ABCDEFG", (18, 74, 20, 12, 11, 11, 14)):
        ws_n.column_dimensions[col].width = w
    ws_n.freeze_panes = "A2"
    ws_n.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Peer comparison sheet ─────────────────────────────────────────────────────
def _build_peer_sheet(wb, peer_df):
    if peer_df is None or peer_df.empty:
        return
    ws_peer = wb.create_sheet("Peer_Comparison")
    for r in dataframe_to_rows(peer_df, index=False, header=True):
        ws_peer.append(r)
    style_header_row(ws_peer, bg=MID_BLUE)
    auto_col_width(ws_peer)
    ws_peer.freeze_panes = "A2"
    ws_peer.auto_filter.ref = f"A1:{get_column_letter(ws_peer.max_column)}1"
    for ri, row in enumerate(ws_peer.iter_rows(min_row=2), 2):
        bg = "D6E4F0" if ri == 2 else (GREY_ROW if ri % 2 == 0 else WHITE)
        for cell in row:
            cell.font   = Font(name="Calibri", size=10, bold=(ri == 2))
            cell.border = _border()
            cell.fill   = PatternFill("solid", fgColor=bg)


# ── Sector comparison sheet ───────────────────────────────────────────────────
def _build_sector_sheet(wb, ticker, df, sector_df):
    if sector_df is None:
        return None
    merged = pd.merge(df[["Date","Cumulative_Index"]], sector_df, on="Date", how="inner")
    merged = merged.rename(columns={"Cumulative_Index": f"{ticker}_Cumulative"})
    ws_s = wb.create_sheet("Sector_Comparison")
    for r in dataframe_to_rows(merged, index=False, header=True):
        ws_s.append(r)
    style_header_row(ws_s)
    auto_col_width(ws_s)
    ws_s.freeze_panes = "A2"
    ws_s.auto_filter.ref = f"A1:{get_column_letter(ws_s.max_column)}1"
    for row in ws_s.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "yyyy-mm-dd" if cell.column == 1 else "0.00"
    return ws_s


# ── Correlation matrix sheet ──────────────────────────────────────────────────
def _build_correlation_sheet(wb, corr_matrix, ticker=None):
    if corr_matrix is None:
        return
    ws_corr = wb.create_sheet("Correlation_Matrix")
    labels  = list(corr_matrix.columns)
    # build_correlation_matrix labels the subject column "Stock" — show the
    # actual ticker in the headers instead of template residue.
    def _disp(lbl):
        return ticker if (lbl == "Stock" and ticker) else lbl
    ws_corr.cell(row=1, column=1, value="Correlation Matrix (Daily Returns)")
    ws_corr.cell(row=1, column=1).font = Font(bold=True, size=12, color=DARK_BLUE, name="Calibri")
    ws_corr.merge_cells(f"A1:{get_column_letter(len(labels)+1)}1")
    for ci, lbl in enumerate(labels, 2):
        _hdr_cell(ws_corr.cell(row=2, column=ci, value=_disp(lbl)), bg=MID_BLUE)
    for ri, lbl in enumerate(labels, 3):
        _hdr_cell(ws_corr.cell(row=ri, column=1, value=_disp(lbl)), bg=MID_BLUE)
        for ci, col_lbl in enumerate(labels, 2):
            val  = corr_matrix.loc[lbl, col_lbl]
            cell = ws_corr.cell(row=ri, column=ci, value=round(float(val), 4))
            cell.number_format = "0.0000"
            cell.font          = Font(name="Calibri", size=10)
            cell.border        = _border()
            cell.alignment     = Alignment(horizontal="center")
    ws_corr.conditional_formatting.add(
        f"B3:{get_column_letter(len(labels)+1)}{len(labels)+2}",
        ColorScaleRule(start_type="num", start_value=-1, start_color="FF9999",
                       mid_type="num",   mid_value=0,    mid_color="FFFFFF",
                       end_type="num",   end_value=1,    end_color="99CCFF"))
    auto_col_width(ws_corr)


# ── Monte Carlo sheet ─────────────────────────────────────────────────────────
def _build_monte_carlo_sheet(wb, mc_sim_df, mc_summary):
    if mc_sim_df is None:
        return None, None, None
    ws_mc = wb.create_sheet("Monte_Carlo")
    ws_mc["A1"] = "Monte Carlo Simulation Summary"
    ws_mc["A1"].font = Font(bold=True, size=13, color=DARK_BLUE, name="Calibri")
    ws_mc["A2"] = "Field"
    ws_mc["B2"] = "Value"
    for cell in ws_mc[2]:
        _hdr_cell(cell, bg=MID_BLUE)
    # _coerce_metric types the value but cannot know what it MEANS, so a price
    # came out as a bare 327.37 and the simulation count as a bare 1000, on a
    # sheet where every other price cell carries a currency format. The key says
    # which is which, and the key is only in scope here.
    _MC_PRICE_KEYS = {"Last Price", "Mean Forecast", "Median (P50)",
                      "Bear Case (P5)", "Low Case (P25)",
                      "Bull Case (P75)", "Best Case (P95)"}
    _MC_COUNT_KEYS = {"Simulations", "Forecast Horizon (days)"}
    for i, (k, v) in enumerate(mc_summary.items(), 3):
        ws_mc.cell(row=i, column=1, value=k).font = Font(name="Calibri", size=10)
        val, vfmt = _coerce_metric(v)
        cv = ws_mc.cell(row=i, column=2, value=val)
        cv.font      = Font(name="Calibri", size=10, bold=True)
        cv.alignment = Alignment(horizontal="right")
        if k in _MC_PRICE_KEYS and isinstance(val, (int, float)):
            cv.number_format = '"$"#,##0.00'
        elif k in _MC_COUNT_KEYS and isinstance(val, (int, float)):
            cv.number_format = '#,##0'
        elif vfmt:
            cv.number_format = vfmt
    summary_end = 3 + len(mc_summary)

    # Percentile paths only. The old sheet dumped 50 raw simulated paths under
    # a header claiming 1,000 simulations — individual paths are noise, they
    # tripled the file size, and the fan chart draws from percentiles anyway.
    start_row_mc = summary_end + 3
    cap = ws_mc.cell(row=start_row_mc - 1, column=1,
                     value=f"Percentile price paths across all {mc_sim_df.shape[1]:,} "
                           f"simulations — the bands the forecast chart draws.")
    cap.font = Font(name="Calibri", size=9, italic=True, color="888888")
    pct_labels = ["P5 (Bear)", "P25 (Low)", "P50 (Median)", "P75 (Bull)", "P95 (Best)"]
    _hdr_cell(ws_mc.cell(row=start_row_mc, column=1, value="Day"), bg=MID_BLUE)
    for j, lbl in enumerate(pct_labels):
        _hdr_cell(ws_mc.cell(row=start_row_mc, column=j + 2, value=lbl), bg=MID_BLUE)
    for day_idx in range(len(mc_sim_df)):
        row_prices = mc_sim_df.iloc[day_idx].values
        r = start_row_mc + 1 + day_idx
        ws_mc.cell(row=r, column=1, value=day_idx)
        for j, pct in enumerate([5, 25, 50, 75, 95]):
            ws_mc.cell(row=r, column=j + 2,
                       value=round(np.percentile(row_prices, pct), 2)).number_format = '"$"#,##0.00'
    for col, w in zip("ABCDEF", (8, 13, 13, 13, 13, 13)):
        ws_mc.column_dimensions[col].width = w
    ws_mc.freeze_panes = f"A{start_row_mc + 1}"
    return ws_mc, start_row_mc, 1


# ── Chart helpers (matplotlib) ────────────────────────────────────────────────
def _mpl_chart(fig):
    """Save a matplotlib figure to a BytesIO PNG buffer."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_style(ax, title, xlabel="Date", ylabel="Price ($)"):
    ax.set_title(title, fontsize=13, fontweight="bold", color="#1F4E79", pad=10)
    ax.set_xlabel(xlabel, fontsize=10, color="#444444")
    ax.set_ylabel(ylabel, fontsize=10, color="#444444")
    ax.tick_params(axis="x", rotation=35, labelsize=8)
    ax.tick_params(axis="y", labelsize=9)
    ax.grid(axis="y", linestyle="--", alpha=0.4, color="#cccccc")
    ax.grid(axis="x", linestyle=":", alpha=0.25, color="#cccccc")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(fontsize=9, framealpha=0.7)


def _add_mpl_image(ws, buf, anchor, width_px=900, height_px=400):
    img = XLImage(buf)
    img.width  = width_px
    img.height = height_px
    ws.add_image(img, anchor)


# ── Charts sheet ──────────────────────────────────────────────────────────────
def _px_ref(ws, col, r0, r1):
    """Reference to one Price_Indicators column, header row included for the name."""
    return Reference(ws, min_col=col, min_row=r0 - 1, max_row=r1)


def _soft(el):
    """Hairline grey outline — for gridlines and axis rules."""
    el.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID_GREY, w=9525))
    return el


def _chart_frame(ch):
    """Shared chrome: no rounded corners, no outer box, legend below the plot.

    Excel defaults put the legend INSIDE the plot area, where it printed on top
    of the data — the MA200 label sat across the MA200 line. Rounded corners and
    a black outer border are the other two tells of an untouched default chart.
    """
    ch.roundedCorners = False
    ch.graphical_properties = GraphicalProperties()
    ch.graphical_properties.line.noFill = True
    if ch.legend is not None:
        ch.legend.position = "b"
        ch.legend.overlay = False
    _soft(ch.x_axis)
    _soft(ch.y_axis)
    return ch


def _line_chart(title, y_title=None, height=8.0, width=22.0):
    ch = LineChart()
    ch.title = title
    ch.style = 2
    ch.height, ch.width = height, width
    ch.y_axis.majorGridlines = _soft(ChartLines())
    ch.x_axis.majorTickMark = "out"
    ch.y_axis.majorTickMark = "out"
    # Excel hides the category axis on some line charts unless told otherwise.
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    # An axis title sits on top of the tick labels at these plot widths, and the
    # chart title already says what is being measured.
    ch.y_axis.title = None
    # Without this Excel colours a single-series chart point-by-point and puts
    # every CATEGORY in the legend — 1,254 dates listed beside a volume chart.
    ch.varyColors = False
    return ch


def _thin_axis(ch, n_points, dates=True):
    """Thin and angle the date labels.

    Left alone, 1,254 categories print as a solid band of rotated text. Thinned
    without angling, Excel finds the horizontal labels still collide and drops
    all but the first, which leaves a chart nobody can orient. Roughly eight
    labels at 45 degrees is what actually fits.
    """
    # Shorten the label before thinning it. The cells hold real dates, so the
    # axis can render "Aug 21" instead of "2021-08-06" — a third of the width,
    # which is what lets Excel fit a useful number of them. tickLblSkip alone
    # did not survive Excel's own auto-fit: it kept the first label and dropped
    # the rest, leaving a chart with no time reference at all.
    # Only when the categories ARE dates. The forecast chart's categories are
    # day numbers, and formatting those as dates rendered day 0 as "Jan 00".
    if dates:
        ch.x_axis.number_format = "mmm yy"
    # No explicit tickLblSkip: Excel overrides it on a category axis this dense
    # (it kept the first label and dropped the other seven). Short labels plus a
    # 45-degree angle let its own auto-fit place a sensible number instead.
    ch.x_axis.textProperties = RichText(
        bodyPr=RichTextProperties(rot=-2700000, vert="horz", anchor="ctr"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=800)))])


def _style_series(ser, hex_colour, width_pt=1.25, dashed=False, name=None):
    if name:
        ser.tx = SeriesLabel(v=name)
    ser.graphicalProperties.line.solidFill = hex_colour
    ser.graphicalProperties.line.width = int(width_pt * 12700)
    if dashed:
        ser.graphicalProperties.line.dashStyle = "dash"
    ser.smooth = False
    ser.marker = Marker(symbol="none")


def _build_charts_sheet(wb, ticker, ws_p, export_df, ws_s, ws_mc_data, full_df=None):
    """Native Excel charts bound to the Price_Indicators cells.

    These were matplotlib PNGs pasted onto the sheet. A picture cannot be
    re-scaled, re-coloured, hovered, or re-pointed at different rows, and it
    goes stale the moment anyone edits the data underneath it — which is the
    whole promise of shipping a workbook rather than a PDF. Every series below
    is a cell range, so the charts move when the data does and behave like
    charts the reader built themselves.
    """
    ws_ch = wb.create_sheet("Charts")
    ws_ch.sheet_view.showGridLines = False
    ws_ch["A1"] = f"{ticker} — Charts"
    ws_ch["A1"].font = Font(size=14, bold=True, color=DARK_BLUE, name="Calibri")
    ws_ch["A2"] = ("Live charts, not images: every series points at the "
                   "Price_Indicators sheet, so edits there redraw these.")
    ws_ch["A2"].font = Font(size=9, italic=True, color="888888", name="Calibri")

    if ws_p is None or ws_p.max_row < 3:
        ws_ch["A4"] = "No price history available to chart."
        return

    # Column letters are resolved from the header row rather than assumed: the
    # price sheet drops columns it has no data for, so positions shift.
    hdr = {ws_p.cell(row=1, column=c).value: c
           for c in range(1, ws_p.max_column + 1)}
    r0, r1 = 2, ws_p.max_row
    cats = Reference(ws_p, min_col=hdr.get("Date", 1), min_row=r0, max_row=r1)
    anchor_row = 4

    def _place(ch):
        nonlocal anchor_row
        ch.set_categories(cats)
        _thin_axis(ch, r1 - r0 + 1)
        _chart_frame(ch)
        ws_ch.add_chart(ch, f"A{anchor_row}")
        anchor_row += 17

    # 1 ── Price with moving averages
    if "Close" in hdr:
        ch = _line_chart(f"{ticker} — Price & Moving Averages", "Price ($)")
        ch.add_data(_px_ref(ws_p, hdr["Close"], r0, r1), titles_from_data=True)
        _style_series(ch.series[-1], "1F4E79", 1.6, name=ticker)
        for name, colour, dash in (("MA50", "70AD47", True),
                                   ("MA200", "C00000", True)):
            if name in hdr:
                ch.add_data(_px_ref(ws_p, hdr[name], r0, r1), titles_from_data=True)
                _style_series(ch.series[-1], colour, 1.1, dashed=dash,
                              name=f"{name[:2]} {name[2:]}")
        _place(ch)

    # 2 ── Cumulative return vs benchmarks
    if "Cumulative_Index" in hdr:
        ch = _line_chart(f"{ticker} — Cumulative Return vs Benchmarks",
                         "Index (100 = start)")
        ch.add_data(_px_ref(ws_p, hdr["Cumulative_Index"], r0, r1),
                    titles_from_data=True)
        _style_series(ch.series[-1], "1F4E79", 1.6, name=ticker)
        for name, colour, label in (("SPY_Cumulative", "C00000", "S&P 500"),
                                    ("QQQ_Cumulative", "00B0F0", "NASDAQ 100")):
            if name in hdr:
                ch.add_data(_px_ref(ws_p, hdr[name], r0, r1), titles_from_data=True)
                _style_series(ch.series[-1], colour, 1.1, dashed=True, name=label)
        _place(ch)

    # 3 ── Volume
    if "Volume" in hdr:
        ch = BarChart()
        ch.type, ch.style = "col", 2
        ch.title = f"{ticker} — Volume"
        ch.height, ch.width = 6.5, 22.0
        ch.gapWidth = 20
        # BarChart doesn't go through _line_chart, so its axes need the same
        # treatment — without this Excel drops the value scale entirely and the
        # bars have no readable magnitude.
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.y_axis.majorGridlines = ChartLines()
        ch.y_axis.numFmt = "#,##0,,\"M\""
        ch.add_data(_px_ref(ws_p, hdr["Volume"], r0, r1), titles_from_data=True)
        ch.series[-1].graphicalProperties.solidFill = "8FAADC"
        ch.series[-1].graphicalProperties.line.noFill = True
        ch.varyColors = False
        ch.legend = None
        ch.y_axis.majorGridlines = _soft(ChartLines())
        _place(ch)

    # 4 ── RSI, with the 70/30 bands drawn as constant series
    if "RSI14" in hdr:
        ch = _line_chart(f"{ticker} — RSI (14)", "RSI", height=6.5)
        ch.add_data(_px_ref(ws_p, hdr["RSI14"], r0, r1), titles_from_data=True)
        _style_series(ch.series[-1], "6C3483", 1.3)
        ch.y_axis.scaling.min, ch.y_axis.scaling.max = 0, 100
        ch.legend = None
        _place(ch)

    # 5 ── Monte Carlo percentile fan, straight off the Monte_Carlo sheet
    if ws_mc_data and ws_mc_data[0]:
        ws_mc, mc_hdr_row, _ = ws_mc_data
        mc_r0, mc_r1 = mc_hdr_row + 1, ws_mc.max_row
        if mc_r1 > mc_r0:
            # Horizon goes in the title, not an axis label: openpyxl gives no
            # control over axis-title placement and Excel printed it straight
            # across the tick numbers.
            _days = mc_r1 - mc_r0
            ch = _line_chart(f"{ticker} — Monte Carlo Forecast "
                             f"({_days} trading days)")
            # P75 and P95 were two greens a reader could not tell apart in a
            # dashed line. Inner band light, outer band dark, red-to-green
            # keeping the bear-to-bull reading.
            for col, colour, wpt in ((2, "C00000", 1.0), (3, "E8A838", 1.0),
                                     (4, "1F4E79", 1.8), (5, "92D050", 1.0),
                                     (6, "375623", 1.0)):
                ch.add_data(Reference(ws_mc, min_col=col, min_row=mc_hdr_row,
                                      max_row=mc_r1), titles_from_data=True)
                _style_series(ch.series[-1], colour, wpt,
                              dashed=(col != 4))
            ch.set_categories(Reference(ws_mc, min_col=1, min_row=mc_r0,
                                        max_row=mc_r1))
            _thin_axis(ch, mc_r1 - mc_r0 + 1, dates=False)
            _chart_frame(ch)
            ws_ch.add_chart(ch, f"A{anchor_row}")
            anchor_row += 17

    ws_ch.column_dimensions["A"].width = 3

# ── Master orchestrator ───────────────────────────────────────────────────────
# ── Fundamentals sheet (EDGAR-sourced statements + quality scores) ────────────
def _build_fundamentals_sheet(wb, fundamentals):
    if not fundamentals or not fundamentals.get("ok"):
        return
    ws = wb.create_sheet("Fundamentals")
    f = fundamentals
    v, m, r, l, g = f["valuation"], f["margins"], f["returns"], f["leverage"], f["growth"]
    q, fc = f.get("quality", {}), f.get("fcf", {})

    def _pctv(x):
        # compute_fundamentals returns percents as 4.9-style numbers; Excel's
        # percent format wants the fraction.
        return x / 100 if isinstance(x, (int, float)) else None

    row = 1
    tc = ws.cell(row=row, column=1,
                 value=f"Fundamentals & Valuation   ·   source: {f.get('source','—')}"
                       f"   ·   FY ending {f.get('as_of','—')}")
    tc.font = Font(size=14, bold=True, color=DARK_BLUE, name="Calibri")
    row += 2

    def section(title, pairs):
        # (label, value, number_format) triples. Values are written as real
        # numbers — the old string cells ("20.55x", "$2.2B") could not be
        # referenced, sorted or charted, and Excel flagged every one.
        nonlocal row
        for col in (1, 2):
            c = ws.cell(row=row, column=col, value=title if col == 1 else None)
            c.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
            c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        row += 1
        for lbl, val, vfmt in pairs:
            cl = ws.cell(row=row, column=1, value=lbl)
            cv = ws.cell(row=row, column=2, value="N/A" if val is None else val)
            cl.font = Font(name="Calibri", size=10)
            cv.font = Font(name="Calibri", size=10, bold=True,
                           color="999999" if val is None else "000000")
            cv.alignment = Alignment(horizontal="right")
            cl.border = cv.border = _border()
            if val is not None and vfmt and isinstance(val, (int, float)):
                cv.number_format = vfmt
            if row % 2 == 0:
                cl.fill = cv.fill = PatternFill("solid", fgColor=GREY_ROW)
            row += 1
        row += 1

    _ig = f.get("implied_growth")
    X   = '0.00"x"'
    BN  = '"$"#,##0.0,,,"B"'
    section("Valuation", [
        ("P/E", v["pe"], X), ("P/S", v["ps"], X), ("P/B", v["pb"], X),
        ("EV / EBITDA", f.get("ev_ebitda"), X),
        ("Earnings Yield", _pctv(v["earnings_yield"]), "0.0%"),
        ("FCF Yield", _pctv(fc.get("fcf_yield")), "0.0%"),
        # Labelled by BASIS, not just as "reverse-DCF implied growth". The
        # Valuation sheet carries a different implied-growth figure — solved
        # against free cash flow at the company's own CAPM WACC — and the two
        # legitimately disagree (net income is not cash, and 9% is not this
        # company's cost of capital). Presenting them under the same name made
        # them look like one number that couldn't make up its mind.
        ("Implied growth — earnings basis, flat 9% discount", _ig, "0.0%"),
        ("  (cross-check only; the headline reverse DCF is on the Valuation sheet)", "", None),
    ])
    section("Profitability & Returns", [
        ("Gross Margin", _pctv(m["gross"]), "0.0%"),
        ("Operating Margin", _pctv(m["operating"]), "0.0%"),
        ("Net Margin", _pctv(m["net"]), "0.0%"),
        ("Return on Equity", _pctv(r["roe"]), "0.0%"),
        ("Return on Assets", _pctv(r["roa"]), "0.0%"),
        ("Free Cash Flow", fc.get("fcf"), BN),
    ])
    section("Growth", [
        ("Revenue YoY", _pctv(g["revenue_yoy"]), "0.0%"),
        ("EPS YoY", _pctv(g["eps_yoy"]), "0.0%"),
        ("Revenue CAGR", _pctv(g["revenue_cagr"]), "0.0%"),
        ("EPS CAGR", _pctv(g["eps_cagr"]), "0.0%"),
    ])
    section("Balance Sheet & Quality", [
        ("Current Ratio", l["current_ratio"], "0.00"),
        ("Debt / Equity", l["debt_to_equity"], "0.00"),
        ("Piotroski F-Score",
         f"{q['f_score']} / 9" if q.get("f_score") is not None else None, None),
        ("Altman Z-Score",
         f"{q['z_score']} ({q['z_zone']})" if q.get("z_score") is not None else None, None),
    ])

    t = f.get("trend", {})
    periods = t.get("periods", [])
    if periods:
        for col in range(1, len(periods) + 2):
            hc = ws.cell(row=row, column=col,
                         value=("Fiscal Period" if col == 1 else periods[col - 2]))
            hc.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
            hc.fill = PatternFill("solid", fgColor=MID_BLUE)
            hc.border = _border()
        row += 1
        for label, key in [("Revenue ($B)", "revenue"), ("Net Income ($B)", "net_income"),
                           ("Free Cash Flow ($B)", "fcf")]:
            ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
            for ci, x in enumerate(t.get(key, []), 2):
                cc = ws.cell(row=row, column=ci,
                             value=(round(x / 1e9, 1) if isinstance(x, (int, float)) else "—"))
                cc.font = Font(name="Calibri", size=10)
                cc.border = _border()
                if isinstance(x, (int, float)):
                    cc.number_format = "0.0"
            ws.cell(row=row, column=1).border = _border()
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 12


# ── Valuation sheet ───────────────────────────────────────────────────────────
# The analytical heart of the workbook. The reverse DCF ("what growth does
# today's price already assume?") leads, and the whole model is written as live
# Excel formulas hanging off a block of editable assumption cells, so a reader
# can change WACC or terminal growth and watch fair value move rather than
# reading a static dump of numbers someone else computed.

# Plain-English translation of analysis.dcf_valuation()'s failure reasons.
_DCF_NO_MODEL_REASONS = {
    "fundamentals unavailable":
        "No company financial statements are published for this security — normal "
        "for ETFs, funds, ADRs and crypto — so there is no cash-flow stream to discount.",
    "no price":
        "No current market price was available, so the model has nothing to anchor to.",
    "no market cap":
        "No market capitalisation was available, so the share count (market cap ÷ "
        "price) could not be derived.",
    "WACC must exceed terminal growth":
        "The discount rate is not above the assumed terminal growth rate, which makes "
        "the terminal value mathematically infinite.",
    "no positive free cash flow to project":
        "This company has not reported positive free cash flow in the years available, "
        "so there is nothing to project forward. A DCF would not be meaningful here — "
        "judge it on the multiples and quality scores on the Fundamentals sheet instead.",
}


def _build_valuation_sheet(wb, ticker, dcf, fundamentals=None):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K"):
        ws.column_dimensions[col].width = 14

    ws.merge_cells("A1:K1")
    ttl = ws["A1"]
    ttl.value = f"{ticker} — Reverse DCF & Fair Value"
    ttl.font  = Font(size=14, bold=True, color=DARK_BLUE, name="Calibri")
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.font      = Font(size=9, italic=True, color="888888", name="Calibri")
    sub.alignment = Alignment(wrap_text=True, vertical="top")

    # ── No credible model: say so in one line rather than shipping a blank tab ─
    if not isinstance(dcf, dict) or not dcf.get("ok"):
        reason = dcf.get("reason", "") if isinstance(dcf, dict) else ""
        sub.value = "A discounted-cash-flow model could not be built for this security."
        ws.row_dimensions[2].height = 16
        _narrative_box(
            ws, 4,
            f"No DCF for {ticker}.  "
            + _DCF_NO_MODEL_REASONS.get(
                reason,
                "The inputs a DCF needs — audited financial statements, a market price "
                "and a positive free-cash-flow history — were not all available.")
            + "  " + DISCLAIMER_SHORT,
            height=64, italic=False, bg="FFF8E1")
        return

    sub.value = ("Two-stage unlevered DCF on free cash flow, run both ways: forwards "
                 "(what is it worth?) and backwards (what does today's price already "
                 "assume?).  Every shaded cell is an input — change one and the "
                 "projection, bridge, fair value and scenarios below all recalculate.")
    ws.row_dimensions[2].height = 26

    # ── Local styling helpers (match the rest of the workbook) ────────────────
    def sec(row, label, span="K"):
        ws.merge_cells(f"A{row}:{span}{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws.row_dimensions[row].height = 18

    def kv2(row, label, value, fmt=None, bold=True, note=None, input_cell=False):
        cl = ws.cell(row=row, column=1, value=label)
        cv = ws.cell(row=row, column=2, value=value)
        cl.font = Font(name="Calibri", size=10)
        cv.font = Font(name="Calibri", size=10, bold=bold)
        cv.alignment = Alignment(horizontal="right")
        cl.border = cv.border = _border()
        if fmt:
            cv.number_format = fmt
        if input_cell:
            cv.fill = PatternFill("solid", fgColor=INPUT_BG)
        if note:
            nc = ws.cell(row=row, column=3, value=note)
            nc.font      = Font(name="Calibri", size=9, italic=True, color="888888")
            nc.alignment = Alignment(vertical="center")
            ws.merge_cells(f"C{row}:K{row}")
        return cv

    proj = dcf.get("projection") or []
    n_yr = len(proj)
    scn  = dcf.get("scenarios") or {}
    g_bear = (scn.get("bear") or {}).get("growth")
    g_bull = (scn.get("bull") or {}).get("growth")
    imp    = dcf.get("market_implied_growth")

    # ── Row plan (fixed up front so every formula can reference its inputs) ────
    R_HEAD_SEC = 4
    R_HEAD_BIG = 5                      # merged 5:6
    R_CONC_SEC = 8
    R_FV, R_PX, R_UP, R_VERD = 9, 10, 11, 12
    R_SCN_SEC, R_SCN_HDR = 14, 15
    R_SCN_BEAR, R_SCN_BASE, R_SCN_BULL, R_SCN_MKT = 16, 17, 18, 19
    R_IN_SEC = 21
    (R_W, R_TG, R_YRS, R_FCF0, R_G1,
     R_GBEAR, R_GBULL, R_ND, R_SH, R_P) = range(22, 32)
    R_PROJ_SEC, R_PROJ_HDR = 33, 34
    R_P0 = 35
    R_PN = R_P0 + max(n_yr - 1, 0)
    R_BR_SEC = R_PN + 2
    R_PVEXP, R_TV, R_PVTV, R_EV, R_NDB, R_EQ = (R_BR_SEC + i for i in range(1, 7))
    R_SENS_SEC = R_EQ + 2
    R_SENS_HDR = R_SENS_SEC + 1
    R_SENS0    = R_SENS_HDR + 1

    # ── 1. The headline: what today's price implies ───────────────────────────
    sec(R_HEAD_SEC, "The Headline — What Today's Price Already Assumes")
    ws.merge_cells(f"A{R_HEAD_BIG}:B{R_HEAD_BIG + 1}")
    big = ws.cell(row=R_HEAD_BIG, column=1,
                  value=(imp if imp is not None else "n/a"))
    if imp is not None:
        big.number_format = FMT_PCT1
    big.font      = Font(name="Calibri", size=30, bold=True, color=DARK_BLUE)
    big.alignment = Alignment(horizontal="center", vertical="center")
    big.fill      = PatternFill("solid", fgColor=TILE_BG)
    big.border    = _border()

    ws.merge_cells(f"C{R_HEAD_BIG}:K{R_HEAD_BIG + 1}")
    if imp is not None:
        head_txt = (
            f"At ${dcf['price']:,.2f} the market is pricing {ticker}'s free cash flow to "
            f"compound at roughly {imp * 100:.1f}% a year for {dcf['years']} years, then "
            f"{dcf['terminal_growth'] * 100:.1f}% in perpetuity, discounted at "
            f"{dcf['wacc'] * 100:.1f}%.\nThe reverse DCF asks one question: is that rate "
            f"plausible for this business?  This model's own base case assumes "
            f"{dcf['base_growth'] * 100:.1f}%.")
    else:
        head_txt = (
            f"Today's ${dcf['price']:,.2f} price sits outside the range of growth rates "
            f"this model can solve for (−20% to +50% a year), so no single implied growth "
            f"rate can be quoted. That is itself informative: the price is not explicable "
            f"by free-cash-flow growth alone at a "
            f"{dcf['wacc'] * 100:.1f}% discount rate.")
    hc = ws.cell(row=R_HEAD_BIG, column=3, value=head_txt)
    hc.font      = Font(name="Calibri", size=11)
    hc.alignment = Alignment(wrap_text=True, vertical="center")
    hc.border    = _border()
    hc.fill      = PatternFill("solid", fgColor=TILE_BG)
    ws.row_dimensions[R_HEAD_BIG].height = 30
    ws.row_dimensions[R_HEAD_BIG + 1].height = 42

    # ── 2. Conclusion — fair value vs price (live off the bridge below) ───────
    sec(R_CONC_SEC, "Conclusion — Fair Value vs Price")
    up = dcf.get("upside")
    # The note must not start with "=" — openpyxl stores it as a formula and
    # Excel renders a visible #NAME? error next to the headline fair value.
    kv2(R_FV, "Base-case fair value / share", f"=B{R_EQ}/B{R_SH}", fmt=FMT_USD,
        note="Equity value ÷ shares, both computed below")
    kv2(R_PX, "Current price", f"=B{R_P}", fmt=FMT_USD)
    kv2(R_UP, "Upside / downside", f'=IF(B{R_PX}=0,"",B{R_FV}/B{R_PX}-1)',
        fmt=FMT_SIGNED)
    verdict, vcol = (("Undervalued vs DCF", GREEN_OK) if up is not None and up > 0.15 else
                     ("Overvalued vs DCF",  RED_BAD)  if up is not None and up < -0.15 else
                     ("Fairly valued vs DCF", DARK_BLUE))
    ws.cell(row=R_VERD, column=1, value="Verdict").font = Font(name="Calibri", size=10)
    vc = ws.cell(row=R_VERD, column=2,
                 value=(f'=IF(B{R_UP}>0.15,"Undervalued vs DCF",'
                        f'IF(B{R_UP}<-0.15,"Overvalued vs DCF","Fairly valued vs DCF"))'))
    vc.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    vc.fill      = PatternFill("solid", fgColor=vcol)
    vc.alignment = Alignment(horizontal="right")
    ws.cell(row=R_VERD, column=1).border = vc.border = _border()
    # "Fairly valued vs DCF" is wider than one 14-char column — unmerged it
    # rendered as "irly valued vs DCF".
    ws.merge_cells(f"B{R_VERD}:C{R_VERD}")

    # ── 3. Scenarios — with the market itself as the fourth row ──────────────
    sec(R_SCN_SEC, "Scenarios — and What the Market Is Assuming")
    for ci, h in enumerate(["Scenario", "Stage-1 FCF growth", "Fair value / share",
                            "Upside / downside"], 1):
        _hdr_cell(ws.cell(row=R_SCN_HDR, column=ci, value=h), bg=MID_BLUE)

    def _scn_row(row, name, growth_ref, fv_formula, highlight=None):
        cells = [
            ws.cell(row=row, column=1, value=name),
            ws.cell(row=row, column=2, value=growth_ref),
            ws.cell(row=row, column=3, value=fv_formula),
            ws.cell(row=row, column=4,
                    value=f'=IF(B{R_P}=0,"",C{row}/B{R_P}-1)'),
        ]
        cells[1].number_format = FMT_PCT1
        cells[2].number_format = FMT_USD
        cells[3].number_format = FMT_SIGNED
        for c in cells:
            c.font   = Font(name="Calibri", size=10, bold=bool(highlight))
            c.border = _border()
            if highlight:
                c.fill = PatternFill("solid", fgColor=highlight)
            c.alignment = Alignment(horizontal="left" if c.column == 1 else "right")

    # Terminal value + PV, in one formula, for a scenario's FCF/PV columns.
    def _scn_fv(fcf_col, pv_col):
        return (f"=((SUM({pv_col}{R_P0}:{pv_col}{R_PN})"
                f"+{fcf_col}{R_PN}*(1+B{R_TG})/(B{R_W}-B{R_TG})/(1+B{R_W})^B{R_YRS})"
                f"-B{R_ND})/B{R_SH}")

    _scn_row(R_SCN_BEAR, "Bear", f"=B{R_GBEAR}", _scn_fv("G", "H"))
    _scn_row(R_SCN_BASE, "Base", f"=B{R_G1}",    f"=B{R_FV}", highlight="D6E4F0")
    _scn_row(R_SCN_BULL, "Bull", f"=B{R_GBULL}", _scn_fv("J", "K"))
    _scn_row(R_SCN_MKT,  "Market (today's price)",
             (imp if imp is not None else "n/a"), f"=B{R_P}", highlight=TILE_BG)
    if imp is None:
        ws.cell(row=R_SCN_MKT, column=2).number_format = "General"

    # ── 4. Model inputs — the editable cells everything else hangs off ────────
    sec(R_IN_SEC, "Model Inputs — every shaded cell is editable")
    kv2(R_W,    "Discount rate (WACC)", dcf["wacc"], fmt=FMT_PCT2, input_cell=True,
        note="Company-specific cost of capital. Raise it and fair value falls.")
    kv2(R_TG,   "Terminal growth rate", dcf["terminal_growth"], fmt=FMT_PCT2,
        input_cell=True, note="Growth forever after the explicit horizon. Must stay below WACC.")
    kv2(R_YRS,  "Explicit forecast horizon (years)", dcf["years"], fmt="0",
        input_cell=True,
        note="Drives discounting and the terminal year; the projection table is fixed "
             f"at {n_yr} rows.")
    kv2(R_FCF0, "Normalised base free cash flow", dcf["base_fcf"], fmt=FMT_BN,
        input_cell=True, note="Mean of the last three positive annual FCF figures.")
    kv2(R_G1,   "Stage-1 FCF growth — base case", dcf["base_growth"], fmt=FMT_PCT1,
        input_cell=True, note="Year-1 growth, fading linearly to the terminal rate.")
    kv2(R_GBEAR, "Stage-1 FCF growth — bear case",
        g_bear if g_bear is not None else dcf["base_growth"], fmt=FMT_PCT1, input_cell=True)
    kv2(R_GBULL, "Stage-1 FCF growth — bull case",
        g_bull if g_bull is not None else dcf["base_growth"], fmt=FMT_PCT1, input_cell=True)
    kv2(R_ND,   "Net debt (total debt − cash)", dcf["net_debt"], fmt=FMT_BN,
        input_cell=True, note="Subtracted from enterprise value to reach equity value.")
    kv2(R_SH,   "Shares outstanding", dcf["shares"], fmt=FMT_SHARES, input_cell=True,
        note="Derived as market cap ÷ price, so the model ties to the quoted price.")
    kv2(R_P,    "Current price", dcf["price"], fmt=FMT_USD, input_cell=True)

    # ── 5. Projection — live formulas, base / bear / bull side by side ────────
    sec(R_PROJ_SEC, "Base-Case Projection  (bear and bull run alongside)")
    for ci, h in enumerate(["Year", "Growth", "Projected FCF", "Discount factor",
                            "PV of FCF", "Bear growth", "Bear FCF", "Bear PV",
                            "Bull growth", "Bull FCF", "Bull PV"], 1):
        _hdr_cell(ws.cell(row=R_PROJ_HDR, column=ci, value=h), bg=MID_BLUE)

    def _fade(g_row, r):
        """Linear fade from the stage-1 rate to terminal, exactly as analysis.py."""
        return (f"=IF($B${R_YRS}<=1,$B${R_TG},$B${g_row}"
                f"+($B${R_TG}-$B${g_row})*(A{r}-1)/($B${R_YRS}-1))")

    for i in range(n_yr):
        r     = R_P0 + i
        first = (i == 0)
        ws.cell(row=r, column=1, value=proj[i]["year"]).number_format = "0"
        ws.cell(row=r, column=2, value=_fade(R_G1, r))
        ws.cell(row=r, column=3,
                value=(f"=$B${R_FCF0}*(1+B{r})" if first else f"=C{r-1}*(1+B{r})"))
        ws.cell(row=r, column=4, value=f"=1/(1+$B${R_W})^A{r}")
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        ws.cell(row=r, column=6, value=_fade(R_GBEAR, r))
        ws.cell(row=r, column=7,
                value=(f"=$B${R_FCF0}*(1+F{r})" if first else f"=G{r-1}*(1+F{r})"))
        ws.cell(row=r, column=8, value=f"=G{r}*D{r}")
        ws.cell(row=r, column=9, value=_fade(R_GBULL, r))
        ws.cell(row=r, column=10,
                value=(f"=$B${R_FCF0}*(1+I{r})" if first else f"=J{r-1}*(1+I{r})"))
        ws.cell(row=r, column=11, value=f"=J{r}*D{r}")
        for ci in range(1, 12):
            c = ws.cell(row=r, column=ci)
            c.font   = Font(name="Calibri", size=10)
            c.border = _border()
            c.alignment = Alignment(horizontal="right")
            if ci in (2, 6, 9):
                c.number_format = FMT_PCT1
            elif ci == 4:
                c.number_format = "0.000"
            elif ci > 1:
                c.number_format = FMT_BN
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GREY_ROW)

    # ── 6. Enterprise value → equity bridge (live) ────────────────────────────
    sec(R_BR_SEC, "Enterprise Value → Equity Bridge")
    for row, lbl, formula in [
        (R_PVEXP, "PV of explicit FCF",  f"=SUM(E{R_P0}:E{R_PN})"),
        (R_TV,    "Terminal value",      f"=C{R_PN}*(1+B{R_TG})/(B{R_W}-B{R_TG})"),
        (R_PVTV,  "PV of terminal value", f"=B{R_TV}/(1+B{R_W})^B{R_YRS}"),
        (R_EV,    "Enterprise value",    f"=B{R_PVEXP}+B{R_PVTV}"),
        (R_NDB,   "Less: net debt",      f"=B{R_ND}"),
        (R_EQ,    "Equity value",        f"=B{R_EV}-B{R_NDB}"),
    ]:
        cv = kv2(row, lbl, formula, fmt=FMT_BN, bold=(row == R_EQ))
        if row == R_EQ:
            cv.fill = PatternFill("solid", fgColor="D6E4F0")
        if row == R_PVTV:
            # How much of the value is the terminal assumption. The standard
            # red flag is >85%: past that the answer is essentially one guess
            # about the far future, and the explicit forecast is decoration.
            # Live, so it moves when the reader changes WACC or growth.
            _tv = ws.cell(row=row, column=3,
                          value=f'=TEXT(B{R_PVTV}/B{R_EV},"0%")&" of enterprise value'
                                f' — above 85% the terminal assumption is doing'
                                f' nearly all the work"')
            _tv.font = Font(name="Calibri", size=9, italic=True, color="888888")

    # ── 7. Sensitivity grid — green above today's price, red below ────────────
    sec(R_SENS_SEC, "Sensitivity — Fair Value / Share  (WACC × Terminal Growth)"
                    "  ·  snapshot computed at generation")
    sens   = dcf.get("sensitivity") or {}
    w_axis = sens.get("wacc_axis") or []
    tg_axis = sens.get("tg_axis") or []
    grid   = sens.get("grid") or []
    if w_axis and tg_axis and grid:
        corner = ws.cell(row=R_SENS_HDR, column=1, value="WACC ╲ Term. g")
        corner.font   = Font(bold=True, size=9, name="Calibri")
        corner.border = _border()
        for cj, tg in enumerate(tg_axis, 2):
            _hdr_cell(ws.cell(row=R_SENS_HDR, column=cj, value=f"{tg * 100:.1f}%"), bg=MID_BLUE)
        for ri, w in enumerate(w_axis):
            r = R_SENS0 + ri
            _hdr_cell(ws.cell(row=r, column=1, value=f"{w * 100:.1f}%"), bg=MID_BLUE)
            for cj, fv in enumerate(grid[ri], 2):
                c = ws.cell(row=r, column=cj, value=(round(fv, 2) if fv else None))
                c.number_format = FMT_USD
                c.font          = Font(name="Calibri", size=10)
                c.border        = _border()
                c.alignment     = Alignment(horizontal="right")
        last_col  = get_column_letter(1 + len(tg_axis))
        grid_rng  = f"B{R_SENS0}:{last_col}{R_SENS0 + len(w_axis) - 1}"
        # Green = the model says it's worth more than the market is charging.
        ws.conditional_formatting.add(grid_rng, CellIsRule(
            operator="greaterThan", formula=[f"$B${R_P}"],
            fill=PatternFill("solid", bgColor=GOOD_FILL),
            font=Font(color=GOOD_TEXT, bold=True)))
        ws.conditional_formatting.add(grid_rng, CellIsRule(
            operator="lessThan", formula=[f"$B${R_P}"],
            fill=PatternFill("solid", bgColor=BAD_FILL),
            font=Font(color=BAD_TEXT)))
        # A static grid sitting beside live formulas is a trap: edit B22 or B23 —
        # the exact thing "every shaded cell is editable" invites — and every
        # other number on the sheet moves while these 25 stay put. Recomputing
        # the full ten-year growth fade inside one array formula per cell would
        # make it live, but it is fragile across Excel versions for what it buys.
        #
        # Instead the sheet notices. This cell is a formula over the same input
        # cells the grid was built from, so the moment they no longer match the
        # axis it was generated at, it says so in the sheet itself rather than
        # leaving the reader to spot it.
        _w_mid  = w_axis[len(w_axis) // 2]
        _tg_mid = tg_axis[len(tg_axis) // 2]
        _stale = ws.cell(
            row=R_SENS0 + len(w_axis),
            column=1,
            value=(f'=IF(AND(ROUND($B${R_W},4)={round(_w_mid, 4)},'
                   f'ROUND($B${R_TG},4)={round(_tg_mid, 4)}),'
                   f'"Grid matches the assumptions above.",'
                   f'"Assumptions changed — this grid was computed at '
                   f'{_w_mid*100:.1f}% WACC / {_tg_mid*100:.1f}% terminal growth '
                   f'and no longer matches the inputs. Regenerate to refresh it.")'))
        _stale.font      = Font(name="Calibri", size=9, italic=True)
        _stale.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=_stale.row, start_column=1,
                       end_row=_stale.row, end_column=1 + len(tg_axis))
        row_after = R_SENS0 + len(w_axis) + 1
    else:
        row_after = R_SENS_HDR + 1

    row_after = _narrative_box(
        ws, row_after,
        f"Each cell is base-case fair value per share at that discount rate and terminal "
        f"growth rate. Green = above today's ${dcf['price']:,.2f} price, red = below. "
        f"These 25 values are computed once when the report is generated; the rest of "
        f"this sheet is live formulas, so editing the inputs above moves everything "
        f"except this grid.",
        height=44, italic=True, bg=TILE_BG)

    row_after = _narrative_box(
        ws, row_after + 1,
        "How to use this sheet: the reverse DCF is the headline — it converts today's "
        "price into the growth rate you would have to believe. Change the discount rate "
        "or the terminal growth rate in the shaded input cells and every projection, "
        "bridge and scenario figure recalculates, so you can test your own assumptions "
        "rather than accept ours. A DCF is a model, not a forecast: base FCF is "
        "normalised over recent years, shares are derived from market cap ÷ price, and "
        "small changes in WACC or terminal growth move fair value materially. "
        + DISCLAIMER_SHORT,
        height=76, italic=True, bg="FFF8E1")

    # ── 8. Football-field chart, driven off the live scenario cells ───────────
    # Anchored at the foot of the sheet so the floating picture can't sit on top
    # of the input notes.
    try:
        ch = BarChart()
        ch.type  = "bar"
        ch.title = "Fair value per share — scenarios vs today's price"
        ch.add_data(Reference(ws, min_col=3, min_row=R_SCN_HDR, max_row=R_SCN_MKT),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=R_SCN_BEAR, max_row=R_SCN_MKT))
        ch.legend = None
        ch.y_axis.numFmt = '$#,##0'
        ch.height, ch.width = 7.0, 15.0
        # Same chrome as the Charts sheet — this one predates it and would
        # otherwise be the only chart in the workbook with rounded corners, a
        # black frame and near-black gridlines.
        ch.varyColors = False
        ch.y_axis.majorGridlines = _soft(ChartLines())
        # Without these Excel drops both scales, leaving bars with no scenario
        # names and no dollar axis — a picture of four rectangles.
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        _chart_frame(ch)
        # Start the chart on a fresh printed page. The model above it is long
        # enough that the chart otherwise straddled a page break, putting three
        # scenario bars on one sheet of paper and the fourth, with the axis, on
        # the next.
        try:
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=row_after))
        except Exception:
            pass
        ws.add_chart(ch, f"A{row_after + 1}")
    except Exception:
        pass


# ── Methodology & Definitions sheet (de-black-boxes every metric) ─────────────
def _build_methodology_sheet(wb, dcf=None):
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 98

    ws.merge_cells("A1:B1")
    ws["A1"].value = "Methodology & Definitions"
    ws["A1"].font  = Font(size=14, bold=True, color=DARK_BLUE, name="Calibri")
    ws.merge_cells("A2:B2")
    ws["A2"].value = ("Every metric in this report, in plain language and with its key "
                      "assumptions — so nothing reads as a black box.")
    ws["A2"].font      = Font(size=9, italic=True, color="888888", name="Calibri")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 4
    rfr = get_risk_free_rate()

    def section(title, items):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws.row_dimensions[row].height = 18
        row += 1
        for term, desc in items:
            a = ws.cell(row=row, column=1, value=term)
            b = ws.cell(row=row, column=2, value=desc)
            a.font = Font(name="Calibri", size=10, bold=True)
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b.font = Font(name="Calibri", size=9)
            b.alignment = Alignment(vertical="top", wrap_text=True)
            a.border = b.border = _border()
            ws.row_dimensions[row].height = max(24, 12.5 * (len(desc) // 108 + 1) + 6)
            row += 1
        row += 1

    section("Risk & Return", [
        ("Annualised Volatility", "Standard deviation of daily returns × √252. How much the price swings; higher = riskier."),
        ("Sharpe Ratio", f"(Annualised return − risk-free rate) ÷ annualised volatility — excess return per unit of total risk. Risk-free rate = current 3-month US Treasury yield via FRED ({rfr*100:.2f}% now). Above 1 is strong."),
        ("Sortino Ratio", "Like Sharpe, but the denominator is downside deviation about zero — √(mean of squared shortfalls below zero, over ALL days) × √252 — so upside volatility isn't penalised. Note this is not the standard deviation of the losing days, which would understate the denominator and flatter the ratio."),
        ("Max Drawdown (60d)", "Deepest fall from a rolling 60-day peak — how bad the worst quarter-ish stretch was. A decline that unfolds over longer than 60 trading days shows up here only in part, so this reads shallower than a full peak-to-trough figure. The Scorecard's risk factor and the PowerPoint deck use the full peak-to-trough drawdown over the whole window."),
    ])
    section("Technical Indicators", [
        ("Moving Average (20/50/200-day)", "Average close over the last N sessions; smooths the trend. Price above the average is bullish, below is cautionary."),
        ("RSI (14)", "Relative Strength Index, 0–100. >70 overbought (stretched up), <30 oversold, 30–70 neutral."),
        ("MACD", "12-day minus 26-day exponential moving average vs a 9-day signal line. MACD above signal = bullish momentum."),
        ("Bollinger Bands (20, 2σ)", "20-day average ± 2 standard deviations. %B shows where price sits within the bands; near the upper band is relatively high."),
        ("Support / Resistance", "Recent levels where the stock repeatedly stalled (resistance) or bounced (support), detected from local highs/lows."),
    ])
    # Read the model's own assumptions rather than restating defaults — WACC is
    # derived per company, so a fixed "9%" in the glossary would be a lie.
    _d      = dcf if isinstance(dcf, dict) and dcf.get("ok") else {}
    _w_txt  = f"{_d['wacc']*100:.1f}%" if _d else "a company-specific discount rate (CAPM-derived WACC)"
    _tg_txt = f"{_d['terminal_growth']*100:.1f}%" if _d else "a long-run terminal growth rate"
    _yr_txt = f"{_d['years']}-year" if _d else "multi-year"

    section("Valuation", [
        ("P/E · P/S · P/B", "Market cap ÷ net income / revenue / book equity — dollars paid per dollar of earnings, sales, or book value."),
        ("EV / EBITDA", "Enterprise value (market cap + debt − cash) ÷ EBITDA — a capital-structure-neutral earnings multiple."),
        ("FCF / Earnings Yield", "Free cash flow / net income ÷ market cap — the cash or earnings return at today's price; higher = cheaper."),
        ("What a Reverse DCF Is",
         "A normal DCF starts with a growth forecast and produces a value. A reverse DCF runs the same model backwards: it takes today's share price as given and solves for the one number that would justify it — the rate at which free cash flow would have to compound over the forecast horizon. "
         "That flips the question from 'what do I think this is worth?' (which needs a forecast you may not have) to 'what would I have to believe to pay this price?' (which you can judge against the company's history, its competitors and its market size). "
         "It needs no analyst estimates, so it cannot inherit their optimism, and it gives you a single falsifiable claim to argue with."),
        ("Reverse-DCF Implied Growth",
         f"The stage-1 free-cash-flow growth rate the current price implies, solved numerically from the same {_yr_txt} two-stage DCF used for fair value ({_w_txt} discount rate, {_tg_txt} terminal growth). "
         "Read it as a hurdle: if the implied rate is comfortably below what the business has actually delivered, the price is undemanding; if it is far above, the price already assumes a lot must go right. "
         "It is blank when today's price falls outside the range of growth rates the model can solve (roughly −20% to +50% a year) — usually a sign the price is being driven by something other than free-cash-flow growth."),
        ("DCF Fair Value",
         f"Two-stage DCF on free cash flow: PV of {_yr_txt} of projected FCF + PV of terminal value − net debt, ÷ shares (market cap ÷ price). "
         f"Discount rate {_w_txt}, terminal growth {_tg_txt}; base FCF normalised over the last three positive annual figures; stage-1 growth fades linearly to the terminal rate. "
         "The Valuation sheet holds the whole model as live Excel formulas over a block of editable assumption cells — change the discount rate, the terminal growth rate or the growth path and every downstream figure recalculates. A model, not a price target."),
        ("Sensitivity Grid",
         "Base-case fair value per share across a 5×5 grid of discount rates and terminal growth rates, on the Valuation sheet. Green cells sit above today's price, red below. "
         "The spread across that grid is the honest measure of how much confidence a DCF deserves for this company."),
    ])
    section("Quality Scores", [
        ("Piotroski F-Score (0–9)", "Nine pass/fail fundamental-quality tests (profitability, leverage, efficiency) across the two latest fiscal years. Higher = higher quality."),
        ("Altman Z-Score", "Bankruptcy-risk score from five ratios. >2.99 'safe', 1.81–2.99 'grey', <1.81 'distress'."),
        ("Technical Posture (0–100)", "A blend of trend (vs 50/200-day MA), RSI, MACD, 52-week location and relative strength. Describes what the indicators say — not a recommendation."),
        ("Stock Scorecard (0–100)", "Weighted composite of seven factors — valuation, growth, profitability, financial health, momentum, risk, sentiment. Describes the stock's profile; not a buy/sell call."),
    ])
    section("Forecasting", [
        ("Monte Carlo Simulation", "Thousands of simulated 1-year price paths (geometric Brownian motion) with drift and volatility from the stock's own daily-return history. P5–P95 are percentiles of simulated ending prices (P5 = only 5% of paths ended lower) — NOT predictions. Assumes log-normal returns and constant volatility; real markets have fat tails and regime shifts. 'Probability of gain' = share of paths ending above today's price."),
        ("Custom Forecast (GARCH + ML)", "Optional variant modelling time-varying volatility (GARCH) and a machine-learned drift before simulating; same percentile interpretation."),
    ])
    section("Sentiment & Data Sources", [
        ("Analyst Consensus", "Wall-Street Buy/Hold/Sell counts aggregated by Finnhub, scored into one verdict. Analysts' view, not QuantWizard's."),
        ("News Relevance & Sentiment", "Headlines are ranked High/Medium/Low for how directly they concern the company (broad round-ups are dropped) and tagged with Polygon's per-article sentiment."),
        ("Data Sources", "Prices: Yahoo Finance (Polygon fallback), split/dividend-adjusted. Fundamentals: SEC EDGAR (Polygon fallback). News & analyst data: Finnhub / Polygon. Risk-free rate: US Treasury via FRED."),
        ("Disclaimer", DISCLAIMER_SHORT + " Figures are generated programmatically for information and education only."),
    ])
    ws.freeze_panes = "A3"


def build_excel(ticker, df, period,
                company_details=None, sector_df=None,
                mc_sim_df=None, mc_summary=None,
                news_list=None, peer_df=None,
                corr_matrix=None,
                resistance_levels=None, support_levels=None,
                summary_text="", bar_size="day", fundamentals=None,
                analyst_data=None, dcf=None):

    wb = Workbook()
    wb.remove(wb.active)

    ws_dash = _build_dashboard(wb, ticker, df, company_details, mc_summary,
                                resistance_levels, support_levels, summary_text,
                                analyst_data=analyst_data, dcf=dcf, fundamentals=fundamentals)
    ws_p, export_df = _build_price_sheet(wb, df, bar_size=bar_size)
    _build_annual_summary(wb, df)
    _build_news_sheet(wb, news_list)
    _build_peer_sheet(wb, peer_df)
    ws_s       = _build_sector_sheet(wb, ticker, df, sector_df)
    _build_correlation_sheet(wb, corr_matrix, ticker)
    ws_mc_data = _build_monte_carlo_sheet(wb, mc_sim_df, mc_summary)
    _build_charts_sheet(wb, ticker, ws_p, export_df, ws_s, ws_mc_data, full_df=df)
    _build_valuation_sheet(wb, ticker, dcf, fundamentals)
    _build_fundamentals_sheet(wb, fundamentals)
    _build_methodology_sheet(wb, dcf=dcf)

    # Final tab order (Cover first, then Dashboard, then the rest). Valuation sits
    # right after the Dashboard — "what is it worth?" follows "what's the answer?".
    # Methodology is the reference appendix at the end. The TOC is built from this
    # same order so the cover links match the tab strip.
    desired = ["Cover","Dashboard","Valuation","Fundamentals","Annual_Summary","Price_Indicators","News_Headlines",
               "Peer_Comparison","Sector_Comparison","Correlation_Matrix",
               "Monte_Carlo","Charts","Methodology"]
    built     = wb.sheetnames                                   # everything except Cover
    toc_order = [s for s in desired if s in built and s != "Cover"]
    extras    = [s for s in built if s not in desired]
    toc_order += extras

    # Cover last so it knows all sheet names; pass df for the KPI band.
    _build_cover(wb, ticker, period, toc_order, df=df)

    # Reorder tabs by sorting the internal sheet list directly. (openpyxl's
    # move_sheet offset is current_index + offset, so the old `index - i`
    # math moved sheets the wrong way and left the order untouched.)
    pos = {name: i for i, name in enumerate(["Cover"] + toc_order)}
    wb._sheets.sort(key=lambda s: pos.get(s.title, 999))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

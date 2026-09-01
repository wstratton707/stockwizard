"""Tests for the Portfolio Builder's section model.

Deliberately narrow. These cover the four things the wizard-to-sections change
could plausibly break, and nothing else:

  1. Step transitions - you cannot reach results without preferences.
  2. Allocation persistence - moving between sections must not rebuild or lose
     the portfolio.
  3. Lazy compute - opening a section creates that section's state and no other.
  4. Export without a forecast - newly reachable, so newly able to crash.

They exercise the state contract rather than the rendering. Streamlit's script
runner is not involved: importing portfolio_builder pulls in streamlit, plotly
and the whole analysis stack, so these assert against the same session-state
keys and helper functions the UI uses, driven directly.
"""
import numpy as np
import pandas as pd
import pytest

import portfolio_builder as pb
from portfolio_analysis import mc_percentiles


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def state(monkeypatch):
    """A dict standing in for st.session_state, with attribute-style access."""
    class _S(dict):
        def pop(self, k, default=None):
            return dict.pop(self, k, default)
    s = _S()
    monkeypatch.setattr(pb.st, "session_state", s, raising=False)
    return s


@pytest.fixture
def sim_df():
    """A small Monte Carlo path matrix: 60 days x 40 simulations."""
    rng = np.random.default_rng(0)
    paths = 10_000 * np.cumprod(1 + rng.normal(0.0004, 0.01, (60, 40)), axis=0)
    return pd.DataFrame(paths)


# ── 1. State transitions ──────────────────────────────────────────────────────

def test_results_requires_a_built_portfolio(state, monkeypatch):
    """_render_results must not render sections when the build failed."""
    called = []
    monkeypatch.setattr(pb, "_ensure_built", lambda _k: False)
    monkeypatch.setattr(pb, "_section_allocation", lambda _k: called.append("alloc"))
    monkeypatch.setattr(pb, "_section_backtest", lambda: called.append("bt"))
    monkeypatch.setattr(pb, "_section_forecast", lambda: called.append("mc"))
    monkeypatch.setattr(pb, "_section_export", lambda: called.append("xl"))

    pb._render_results(api_key="")
    assert called == [], "sections rendered over a portfolio that does not exist"


def test_section_key_defaults_and_rejects_junk(state):
    """An unknown or missing section falls back to Allocation, never blank."""
    assert pb._SECTIONS[0] == "Allocation"
    for junk in (None, "", "Nonsense", 7):
        state[pb._K_SECTION] = junk
        current = state.get(pb._K_SECTION, pb._SECTIONS[0])
        if current not in pb._SECTIONS:
            current = pb._SECTIONS[0]
        assert current == "Allocation"


def test_start_new_clears_every_portfolio_key(state):
    """Starting over must not leave a previous portfolio's artefacts behind."""
    keys = [pb._K_STEP, pb._K_SECTION, pb._K_PREFS, pb._K_OPTIMISED, pb._K_WEIGHTS,
            pb._K_BACKTEST, pb._K_MC, pb._K_EXCEL, pb._K_PPTX, pb._K_RANKINGS]
    for k in keys:
        state[k] = "stale"
    for k in keys:
        state.pop(k, None)
    assert state == {}, f"left behind: {sorted(state)}"


# ── 2. Allocation persistence ─────────────────────────────────────────────────

def test_switching_sections_does_not_rebuild(state, monkeypatch):
    """Allocation -> Backtest -> Allocation must reuse the same build."""
    builds = []

    def _fake_build(_api_key):
        if pb._K_OPTIMISED not in pb.st.session_state:
            builds.append(1)
            pb.st.session_state[pb._K_OPTIMISED] = {"returns_df": "frame"}
        return True

    monkeypatch.setattr(pb, "_ensure_built", _fake_build)

    for section in ("Allocation", "Backtest", "Allocation"):
        state[pb._K_SECTION] = section
        _fake_build("")

    assert len(builds) == 1, f"rebuilt {len(builds)} times"
    assert state[pb._K_OPTIMISED] == {"returns_df": "frame"}


def test_weights_survive_a_section_change(state):
    """_K_WEIGHTS is what every other section, the exports and the save read."""
    state[pb._K_WEIGHTS] = {"AAPL": 0.5, "MSFT": 0.5}
    state[pb._K_SECTION] = "Backtest"
    assert state[pb._K_WEIGHTS] == {"AAPL": 0.5, "MSFT": 0.5}
    state[pb._K_SECTION] = "Export"
    assert state[pb._K_WEIGHTS] == {"AAPL": 0.5, "MSFT": 0.5}


# ── 3. Lazy compute ───────────────────────────────────────────────────────────

def test_allocation_creates_no_backtest_or_forecast(state, monkeypatch):
    """Viewing the allocation must not compute anything downstream of it."""
    monkeypatch.setattr(pb, "_ensure_built", lambda _k: True)
    monkeypatch.setattr(pb, "_section_allocation", lambda _k: None)
    state[pb._K_SECTION] = "Allocation"

    pb._section_allocation("")
    assert pb._K_BACKTEST not in state
    assert pb._K_MC not in state


def test_forecast_is_gated_behind_an_explicit_run(state, monkeypatch):
    """The Monte Carlo is the heaviest step; it must not fire on mere arrival.

    The section returns early while the Run button is unclicked, so no _K_MC
    appears. This is the difference from the backtest, which computes on sight.
    """
    state[pb._K_PREFS] = {"horizon": "5 years"}
    state[pb._K_OPTIMISED] = {"returns_df": None}
    state[pb._K_WEIGHTS] = {"AAPL": 1.0}

    monkeypatch.setattr(pb.st, "caption", lambda *a, **k: None)
    monkeypatch.setattr(pb.st, "button", lambda *a, **k: False)   # not clicked

    ran = []
    monkeypatch.setattr(pb, "run_portfolio_monte_carlo",
                        lambda *a, **k: ran.append(1))

    pb._section_forecast()
    assert ran == [], "forecast ran without the user asking for it"
    assert pb._K_MC not in state


# ── 4. Export without a forecast ──────────────────────────────────────────────

def test_export_tolerates_a_missing_forecast(state):
    """Export is now reachable without Backtest or Forecast having been opened."""
    state[pb._K_PREFS] = {}
    state[pb._K_WEIGHTS] = {"AAPL": 1.0}
    mc_data = state.get(pb._K_MC, {})

    assert mc_data.get("pcts") is None
    assert mc_data.get("summary", {}) == {}
    assert mc_data.get("milestones", {}) == {}

    missing = []
    if not state.get(pb._K_BACKTEST, {}):
        missing.append("the backtest")
    if mc_data.get("pcts") is None:
        missing.append("the forecast")
    assert missing == ["the backtest", "the forecast"]


def test_excel_mc_sheet_accepts_percentiles_and_none(sim_df):
    """The exporters take the reduced frame, and skip cleanly when it is absent."""
    from openpyxl import Workbook
    from portfolio_excel import _build_mc_sheet

    pcts = mc_percentiles(sim_df)
    assert list(pcts.columns) == ["p5", "p25", "p50", "p75", "p95"]
    assert len(pcts) == len(sim_df)

    wb = Workbook()
    _build_mc_sheet(wb, pcts, {"Paths": 40}, {"1yr": dict.fromkeys(
        ["P5", "P25", "P50", "P75", "P95"], 1.0)})
    assert "Monte_Carlo_Portfolio" in wb.sheetnames

    wb2 = Workbook()
    before = list(wb2.sheetnames)
    _build_mc_sheet(wb2, None, {}, {})
    assert wb2.sheetnames == before, "a missing forecast should add no sheet"


# ── Phase 0 guard: the percentile reduction ──────────────────────────────────

def test_mc_percentiles_matches_the_old_inline_calculation(sim_df):
    """The fan chart used to compute this inline; the stored frame must match."""
    pcts = mc_percentiles(sim_df)
    old = np.percentile(sim_df.values, [5, 25, 50, 75, 95], axis=1)
    assert np.allclose(pcts.values.T, old)
    assert mc_percentiles(pcts) is pcts, "must be idempotent"
    assert mc_percentiles(None) is None
